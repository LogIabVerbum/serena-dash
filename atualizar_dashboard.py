"""
WMS Serena - Rotina diária de atualização do Dashboard
======================================================
Executa todo dia às 10h (via Agendador de Tarefas do Windows)

O que faz:
  1. Faz backup do Quadro de Envios na pasta do mês (062026, 072026, etc.)
  2. Lê o arquivo do Sankhya e integra novos lançamentos no Quadro de Envios
  3. Processa os dados e atualiza o dashboard_logistica.html

Uso manual:
    python atualizar_dashboard.py
    ou dois cliques no ATUALIZAR_DASHBOARD.bat
"""

import sys, json, re, unicodedata, shutil

# Forcar UTF-8 no terminal (compativel com Windows cp1252)
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from pathlib import Path
from datetime import datetime

# -- Caminhos ------------------------------------------------------------------
# Pasta onde estão os arquivos originais (Quadro de Envios, correlação, etc.)
PASTA_RAIZ = Path(r"C:\Users\a2p2t\OneDrive - Instituto Alfa e Beto\Logística-IAB\1 - QUADRO DE ENVIOS\Rotina_Preenchimento")

# Pasta do GitHub Desktop (onde o HTML publicado fica)
PASTA_GITHUB = Path(__file__).parent
# Backup: detecta automaticamente qual maquina esta rodando
# Adicione novos caminhos conforme necessario
_BACKUPS_CONHECIDOS = [
    Path(r"C:\Users\a2p2t\OneDrive - Instituto Alfa e Beto\Logistica-IAB\1 - QUADRO DE ENVIOS\2026"),
    Path(r"C:\Users\mique\OneDrive - Instituto Alfa e Beto\Logistica-IAB\1 - QUADRO DE ENVIOS\2026"),
    Path(r"C:\Users\AugustoCesar\OneDrive - Instituto Alfa e Beto\Arquivos de Anderson Pissaia - Logistica-IAB\1 - QUADRO DE ENVIOS\2026"),
]
_backup_found = next((p for p in _BACKUPS_CONHECIDOS if p.parent.exists()), None)
PASTA_BACKUPS = _backup_found if _backup_found else PASTA_RAIZ / 'backups'
JSON_DASHBOARD  = PASTA_RAIZ / 'dados_dashboard.json'
HTML_DASHBOARD  = PASTA_GITHUB / 'dashboard_logistica.html'
HTML_NEUTRO     = PASTA_GITHUB / 'dashboard_neutro.html'
MUNICIPIOS_JSON = PASTA_RAIZ / 'municipios_ibge.json'
CORRELACAO_FILE = PASTA_RAIZ / 'CORELACAO_NOME_TRANSP.xlsx'

# Nome do arquivo Sankhya base - aceita .xlsx ou .xls
_s1 = PASTA_RAIZ / 'arquivo_base.xlsx'
_s2 = PASTA_RAIZ / 'arquivo_base.xls'
_s3 = PASTA_RAIZ / 'arquivo.xlsx'
_s4 = PASTA_RAIZ / 'arquivo.xls'
ARQUIVO_SANKHYA = next((f for f in [_s1,_s2,_s3,_s4] if f.exists()), _s1)

# Quadro de Envios - aceita com underline ou espaço
_q1 = PASTA_RAIZ / 'Quadro_de_Envios.xlsx'
_q2 = PASTA_RAIZ / 'Quadro de Envios.xlsx'
QUADRO_ENVIOS = _q1 if _q1.exists() else _q2

# Cadastro consolidado de clientes
_c1 = PASTA_RAIZ / 'cadastro_clientes_consolidado.csv'
_c2 = PASTA_RAIZ / 'historico' / 'cadastro_clientes_consolidado.csv'
CADASTRO_CSV = next((f for f in [_c1, _c2] if f.exists()), _c1)

# Arquivos históricos
_hd = PASTA_RAIZ / 'historico'
HIST_2012 = next((p for p in [_hd/'Historico_2012_a_2015.xlsx', PASTA_RAIZ/'Historico_2012_a_2015.xlsx'] if p.exists()), None)
HIST_2016 = next((p for p in [_hd/'Historico_2016_a_2019.xlsx', PASTA_RAIZ/'Historico_2016_a_2019.xlsx'] if p.exists()), None)
HIST_2020 = next((p for p in [_hd/'Historico_2020_a_2024.xlsx', PASTA_RAIZ/'Historico_2020_a_2024.xlsx'] if p.exists()), None)

def log(msg, tipo='info'):
    icone = {'info':'[ARQ]','ok':'[OK]','erro':'[ERRO]','aviso':'[AVISO]','proc':'[...]'}.get(tipo,'>')
    print(f"  {icone} {msg}")

# -- TABELA DE CORRELAÇÃO DE TRANSPORTADORAS -----------------------------------
def carregar_correlacao():
    """
    Carrega CORELACAO_NOME_TRANSP.xlsx e retorna dict {nome_sistema_upper: nome_dashboard}.
    Inclui mapeamentos fixos para variações conhecidas (códigos numéricos, Title Case, etc.)
    """
    import pandas as pd

    # Mapeamentos extras fixos (variações conhecidas não cobertas pela tabela)
    extras = {
        '492 - TG TRANSPORTES GERAIS E DISTRIBUICAO DF':             'TG',
        '232 - JADLOG LOGISTICA S.A':                                'JADLOG',
        '115 - BRASPRESS TRANSPORTES URGENTES LTDA':                 'BRASPRESS',
        '127 - FERNANDES E GARCIA AGENCIA DE SERVICOS POSTAIS LTDA': 'CORREIOS',
        '118 - EMPRESA BRASILEIRA DE CORREIOS E TELEGRAFOS':         'CORREIOS',
        '301 - BTU - BRASPRESS - UDI':                               'BRASPRESS',
        'BTU - BRASPRESS - UDI':                                     'BRASPRESS',
        'PRÓPRIO':                                                   'PRÓPRIO',
        '306 - TG TRANSPORTES GERAIS E DISTRIBUICAO LTDA':           'TG',
        'CLARA':                                                     'CLARA',
        'CLARA TRANSPORTES':                                         'CLARA',
        '126 - AZUL LINHAS AEREAS BRASILEIRAS S.A.':                 'AZUL CARGO',
        '120 - FEDEX':                                               'TNT',
        'ATIVA':                                                     'ATIVA',
        '521 - CLARA TRANSPORTES E OPERACOES LOGISTICA LTDA.':       'CLARA',
        '116 - UBERLOG TRANSPORTES E SERVICOS LTDA':                 'UBERLOG',
        '629 - CLARA TRANSPORTES E OPERACOES LOGISTICA LTDA. UDI':   'CLARA',
        '321 - BRASPRESS TRANSPORTES URGENTES LTDA':                 'BRASPRESS',
        '986 - JAMEF ENCOMENDAS URGENTES':                           'JAMEF',
        'AZUL':                                                      'AZUL CARGO',
        '1013 - LIDER ENTREGAS INTELIGENTES':                        'LIDER',
        'JADLOG':                                                     'JADLOG',
        '1083 - JAMEF ENCOMENDAS URGENTES':                          'JAMEF',
        '569 - BRASPRESS TRANSPORTES URGENTES LTDA':                 'BRASPRESS',
        '1241 - ATIVA DISTRIBUICAO E LOGISTICA LTDA':                'ATIVA',
        '332 - GOL LINHAS AEREAS S.A':                               'GOLLOG',
        '1563 - PANSERVICE TRANSPORTES E LOCACAO LTDA':              'PANSERVICE',
        '1583 - ATIVA DISTRIBUICAO E LOGISTICA LTDA':                'ATIVA',
        '1682 - ATIVA DISTRIBUICAO E LOGISTICA LTDA':                'ATIVA',
        '1685 - ATIVA DISTRIBUICAO E LOGISTICA LTDA':                'ATIVA',
        'CW3':                                                       'CW3',
        'TG TRANSPORTES':                                            'TG',
        'PASSARO TRANSPORTES':                                       'PASSARO',
        'ALLI':                                                      'ALLI',
        'FERNANDES E GARCIA AGENCIA DE SERVICOS POSTAIS LTDA':       'CORREIOS',
        'EMPRESA BRASILEIRA DE CORREIOS E TELEGRAFOS':               'CORREIOS',
        'NAO INFORMADO':                                             'SEM PARCEIRO',
        'NÃO INFORMADO':                                             'SEM PARCEIRO',
        '492 - TG TRANSPORTES GERAIS E DISTRIBUICAO DF':            'TG',
    }

    cor_map = {k.strip().upper(): v for k, v in extras.items()}

    if CORRELACAO_FILE.exists():
        try:
            df_cor = pd.read_excel(CORRELACAO_FILE, engine='openpyxl')
            for _, row in df_cor.iterrows():
                key = str(row['NOME NO SISTEMA']).strip().upper()
                val = str(row['NOME NO DASHIBOARD']).strip()
                cor_map[key] = val
            log(f"Correlação carregada: {len(cor_map)} entradas", 'ok')
        except Exception as e:
            log(f"Erro ao carregar correlação: {e}", 'aviso')
    else:
        log(f"CORELACAO_NOME_TRANSP.xlsx não encontrada - usando mapeamentos fixos", 'aviso')

    return cor_map

def map_transp(nome, cor_map):
    """Mapeia nome bruto da transportadora para nome padronizado do dashboard."""
    import pandas as pd
    if pd.isna(nome) or str(nome).strip() == '':
        return 'SEM PARCEIRO'
    key = str(nome).strip().upper()
    resultado = cor_map.get(key)
    if resultado:
        return resultado
    # Avisa no log mas não interrompe
    log(f"Transportadora sem mapeamento: {repr(nome)} - mantendo nome original", 'aviso')
    return str(nome).strip()

# -- CADASTRO DE CLIENTES ------------------------------------------------------
def carregar_cadastro():
    if not CADASTRO_CSV.exists():
        log(f"Cadastro não encontrado: {CADASTRO_CSV.name}", 'aviso')
        return {}
    try:
        import pandas as pd
        df = pd.read_csv(CADASTRO_CSV, dtype=str, encoding='utf-8-sig')
        df['_norm'] = df['CNPJ_CPF'].apply(lambda x: ''.join(filter(str.isdigit, str(x))))
        lookup = dict(zip(df['_norm'], df['Tipo_Cliente'].fillna('')))
        log(f"Cadastro carregado: {len(lookup):,} clientes", 'ok')
        return lookup
    except Exception as e:
        log(f"Erro ao carregar cadastro: {e}", 'erro')
        return {}

def tipo_cliente_lookup(cnpj, lookup):
    if not lookup:
        return ''
    norm = ''.join(filter(str.isdigit, str(cnpj)))
    return lookup.get(norm, '')

# -- 1. BACKUP -----------------------------------------------------------------
def fazer_backup():
    print("\n[1/3] BACKUP DO QUADRO DE ENVIOS")
    if not QUADRO_ENVIOS.exists():
        log(f"Quadro de Envios não encontrado: {QUADRO_ENVIOS}", 'erro')
        return False

    hoje      = datetime.now()
    ano_pasta = hoje.strftime('%Y')
    mes_pasta = hoje.strftime('%m%Y')
    nome_bkp  = f"Quadro_de_Envios_{hoje.strftime('%d-%m-%Y')}.xlsx"
    # Estrutura: backups/2026/062026/
    pasta_mes = PASTA_BACKUPS / ano_pasta / mes_pasta
    pasta_mes.mkdir(parents=True, exist_ok=True)
    destino   = pasta_mes / nome_bkp

    shutil.copy2(QUADRO_ENVIOS, destino)
    log(f"Backup salvo: backups/{ano_pasta}/{mes_pasta}/{nome_bkp}", 'ok')
    return True


# Mapa UF -> Região
_UF_REGIAO = {
    'AC':'NORTE','AM':'NORTE','AP':'NORTE','PA':'NORTE','RO':'NORTE','RR':'NORTE','TO':'NORTE',
    'AL':'NORDESTE','BA':'NORDESTE','CE':'NORDESTE','MA':'NORDESTE','PB':'NORDESTE',
    'PE':'NORDESTE','PI':'NORDESTE','RN':'NORDESTE','SE':'NORDESTE',
    'DF':'CENTRO-OESTE','GO':'CENTRO-OESTE','MS':'CENTRO-OESTE','MT':'CENTRO-OESTE',
    'ES':'SUDESTE','MG':'SUDESTE','RJ':'SUDESTE','SP':'SUDESTE',
    'PR':'SUL','RS':'SUL','SC':'SUL',
}
def get_regiao(uf):
    return _UF_REGIAO.get(str(uf).strip().upper(), '')

# -- 2. INTEGRAÇÃO SANKHYA -----------------------------------------------------
def integrar_sankhya():
    print("\n[2/3] INTEGRAÇÃO SANKHYA -> QUADRO DE ENVIOS")

    if not ARQUIVO_SANKHYA.exists():
        log(f"Arquivo Sankhya não encontrado: {ARQUIVO_SANKHYA.name}", 'aviso')
        log("Pulando integração - continuando só com o Quadro de Envios atual.", 'aviso')
        return True

    try:
        import pandas as pd
    except ImportError:
        log("pandas não instalado. Execute: pip install pandas openpyxl xlrd", 'erro')
        return False

    log(f"Lendo: {ARQUIVO_SANKHYA.name}")
    try:
        df_s = pd.read_excel(ARQUIVO_SANKHYA, header=2)
        df_s.columns = [str(c).strip() for c in df_s.columns]
        df_s = df_s[pd.to_numeric(df_s['NRO_NFE'], errors='coerce').notna()].copy()
        df_s['NRO_NFE'] = df_s['NRO_NFE'].astype(int).astype(str)
        log(f"Sankhya: {len(df_s)} registros válidos")
    except Exception as e:
        log(f"Erro ao ler arquivo_base: {e}", 'erro')
        return False

    COL_NF        = 'NRO_NFE'
    COL_TRANSP    = 'TRANSPORTADORA'
    cols_sankhya  = set(df_s.columns)

    if COL_NF not in cols_sankhya:
        log(f"Coluna '{COL_NF}' não encontrada no arquivo_base", 'erro')
        return False

    log(f"Lendo Quadro de Envios: {QUADRO_ENVIOS.name}")
    try:
        df_q = pd.read_excel(QUADRO_ENVIOS, sheet_name='Lançamentos')
        # Chave NF + CNPJ_DEST + DATA para identificar notas únicas
        # Não usa transportadora pois pode ter nome diferente entre Sankhya e Quadro
        import re as _re
        def _norm_cnpj(c):
            return _re.sub(r'[.\-/\s,]', '', str(c)).strip().lstrip('0') if c else ''

        df_q['_NF_str']   = pd.to_numeric(df_q['NF'], errors='coerce').dropna().astype(int).astype(str)
        df_q['_CNPJ_str'] = df_q['CNPJ / CPF'].astype(str).str.strip()
        # Detectar nome da coluna de data automaticamente
        _col_data_q = next((c for c in df_q.columns if 'coleta' in c.lower()), 'Data Coleta')
        df_q['_DATA_str'] = pd.to_datetime(df_q[_col_data_q], errors='coerce').dt.strftime('%Y-%m-%d')
        df_q['_CNPJ_N']   = df_q['_CNPJ_str'].apply(_norm_cnpj)
        # Chave completa: NF + CNPJ_norm + Data
        chaves_quadro = set(df_q['_NF_str'] + '|' + df_q['_CNPJ_N'] + '|' + df_q['_DATA_str'])
        # Chave sem data: NF + CNPJ — usada para evitar duplicatas mesmo com data diferente
        chaves_quadro_sem_data = set(df_q['_NF_str'] + '|' + df_q['_CNPJ_N'])
        nfs_quadro    = set(df_q['_NF_str'])
        log(f"Quadro: {len(df_q)} lançamentos existentes")
    except Exception as e:
        log(f"Erro ao ler Quadro de Envios: {e}", 'erro')
        return False

    # NF+CNPJ não existente no Quadro = nova
    df_s['_CNPJ_N']    = df_s['CPF_CNPJ_DESTINATARIO'].astype(str).apply(_norm_cnpj)
    df_s['_DATA_str']  = pd.to_datetime(df_s['DT_EMISSAO_NF'], errors='coerce').dt.strftime('%Y-%m-%d')
    df_s['_chave']     = df_s[COL_NF] + '|' + df_s['_CNPJ_N'] + '|' + df_s['_DATA_str']
    df_s['_chave_sd']  = df_s[COL_NF] + '|' + df_s['_CNPJ_N']  # sem data
    # NF é nova se NÃO existe no quadro nem pela chave completa NEM pela chave sem data
    # Isso evita duplicatas quando a mesma NF+CNPJ já existe com data diferente
    nfs_realmente_novas = set(
        df_s[~df_s['_chave'].isin(chaves_quadro) & ~df_s['_chave_sd'].isin(chaves_quadro_sem_data)][COL_NF]
    )

    # NFs incompletas: existem no Quadro (pela chave) mas sem Cons Analise
    nfs_incompletas = set()
    mask_incompleta = (
        (df_q['_NF_str'] + '|' + df_q['_CNPJ_N'] + '|' + df_q['_DATA_str']).isin(
            set(df_s['_chave'])
        ) &
        (df_q['Cons Analise'].isna() | (df_q['Cons Analise'].astype(str).str.strip() == ''))
    )
    if mask_incompleta.any():
        nfs_incompletas = set(
            pd.to_numeric(df_q[mask_incompleta]['NF'], errors='coerce')
            .dropna().astype(int).astype(str)
        )

    nfs_novas = nfs_realmente_novas | nfs_incompletas
    log(f"NFs realmente novas: {len(nfs_realmente_novas)} | Incompletas a corrigir: {len(nfs_incompletas)}")
    if nfs_novas:
        log(f"NFs a processar: {sorted(list(nfs_novas))[:15]}", 'info')

    if len(nfs_novas) == 0:
        log("Nenhuma NF nova - Quadro já está atualizado!", 'ok')
        return True

    # Para NFs incompletas: remover as linhas antigas antes de reinserir
    if nfs_incompletas:
        linhas_remover = df_q[mask_incompleta].index.tolist()
        log(f"Removendo {len(linhas_remover)} linhas incompletas para reinserção...", 'aviso')
        # Reabrir e salvar sem essas linhas
        from openpyxl import load_workbook as _lw
        _wb = _lw(QUADRO_ENVIOS)
        _ws = _wb['Lançamentos']
        # Deletar de baixo para cima para não deslocar índices
        for idx_rem in sorted(linhas_remover, reverse=True):
            _ws.delete_rows(idx_rem + 2)  # +2: linha 1=header, pandas começa em 0
        _wb.save(QUADRO_ENVIOS)
        log(f"Linhas incompletas removidas. Reinserindo com dados completos...", 'ok')
        # Recarregar Quadro após remoção
        df_q = pd.read_excel(QUADRO_ENVIOS, sheet_name='Lançamentos')
        nfs_quadro = set(
            pd.to_numeric(df_q['NF'], errors='coerce')
            .dropna().astype(int).astype(str)
        )

    tem_transp = COL_TRANSP in cols_sankhya
    if not tem_transp:
        log(f"Campo '{COL_TRANSP}' não disponível no Sankhya.", 'aviso')
        log(f"{len(nfs_novas)} NFs novas detectadas - adicione manualmente no Quadro.", 'aviso')
        log(f"NFs novas: {sorted(list(nfs_novas))[:10]}{'...' if len(nfs_novas)>10 else ''}", 'aviso')
        return True

    log(f"Integrando {len(nfs_novas)} NFs novas...", 'ok')
    cadastro_lookup = carregar_cadastro()

    CNPJ_EMP = {'08458084000113': 'IAB', '57638518000253': 'VERBUM'}
    def get_empresa(cnpj_rem):
        cnpj = str(cnpj_rem).replace('.','').replace('/','').replace('-','').strip()
        return CNPJ_EMP.get(cnpj, 'IAB')

    df_novas = df_s[df_s[COL_NF].isin(nfs_novas)].copy()
    df_novas['DT_EMISSAO_NF'] = pd.to_datetime(df_novas['DT_EMISSAO_NF'], errors='coerce')

    from openpyxl import load_workbook
    from openpyxl.styles import Font

    wb = load_workbook(QUADRO_ENVIOS)
    ws = wb['Lançamentos']

    # -- Mapear cabeçalho linha 1 -> número de coluna ---------------------------
    col_map = {}
    # Aliases para campos cujo nome pode variar no arquivo
    _ALIASES = {
        'Data Coleta'  : lambda n: 'coleta' in n.lower(),
        'Nº Pedido'    : lambda n: 'pedido' in n.lower(),
        'No Pedido'    : lambda n: 'pedido' in n.lower(),
        'Difal'        : lambda n: 'ifal' in n.lower(),
    }
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is not None:
            nome = str(val).strip()
            col_map[nome] = col_idx
            # Registrar aliases automaticamente
            for alias, fn in _ALIASES.items():
                if fn(nome) and alias not in col_map:
                    col_map[alias] = col_idx
    log(f"Colunas mapeadas no cabeçalho: {len(col_map)}", 'ok')

    # Campos com fórmula - apenas os que NÃO vêm do arquivo_base
    # Cliente, Municipio, UF, Região agora vêm direto do Sankhya
    CAMPOS_FORMULA = {
        'Ano', 'Mês',                           # =YEAR(A) / =MONTH(A) - baseado na Data Coleta
        # 'País' removido - preenchido diretamente como 'BRASIL'
        'Cep',                                   # não disponível no Sankhya
        'Total Impostos',                        # =SUM(ICMS+Difal+FCP)
        'Custo Embalagem',                       # operador preenche
        'Prev. Envio',                           # calculado pelo Excel
        'Prev. Entrega Cliente',                 # calculado pelo Excel
        'Prev. Entrega Transporadora',           # calculado pelo Excel
        'Entrega No Prazo',                      # calculado pelo Excel
        'Entrega em Atraso',                     # calculado pelo Excel
        'Total Entregas',                        # calculado pelo Excel
        'Origem_Escola',                         # calculado pelo Excel
    }

    # -- Guardar fórmulas da última linha com dados reais ----------------------
    # Serão replicadas (com no de linha ajustado) para cada nova linha inserida
    ultima_linha_ref = ws.max_row
    formulas_ref = {}
    import re as _re
    for col_idx in range(1, ws.max_column + 1):
        cell_ref = ws.cell(row=ultima_linha_ref, column=col_idx)
        if isinstance(cell_ref.value, str) and cell_ref.value.startswith('='):
            formulas_ref[col_idx] = cell_ref.value  # fórmula da linha de referência

    log(f"Fórmulas capturadas para replicação: {len(formulas_ref)}", 'ok')
    prox_linha = ws.max_row + 1

    for _, r in df_novas.iterrows():
        data    = r['DT_EMISSAO_NF']
        empresa = get_empresa(r.get('CNPJ_REMETENTE',''))

        # -- Ajuste 1: Data Coleta sem hora (só date) --------------------------
        data_coleta = data.date() if pd.notna(data) else None

        # -- Ajuste 2: Ano Ref na coluna correta (AQ = col 43) -----------------
        ano_ref = data.year if pd.notna(data) else ''

        # -- Ajuste 3: Região automática pela UF -------------------------------
        uf_dest = str(r.get('UF_DESTINO','')).strip().upper()
        regiao  = get_regiao(uf_dest)

        # -- Ajuste 4: Natureza da operação - notas de entrada sem valor negativo
        natureza_raw = str(r.get('TIPO_OPERACAO','')).strip()
        vr_nf_raw    = r.get('VALOR_TOTAL_NOTA', 0)
        vr_nf        = abs(float(vr_nf_raw)) if vr_nf_raw else 0  # sempre positivo

        # -- Ajuste 5: Transportadora em branco gera destaque para o operador --
        # Transportadora: NAO INFORMADO = vazio (operador preenche no Quadro)
        _transp_raw = str(r.get('TRANSPORTADORA','')).strip()
        _nao_info   = {'NAO INFORMADO','NÃO INFORMADO','NAO_INFORMADO','','NONE','NAN'}
        transp = '' if _transp_raw.upper() in _nao_info else _transp_raw

        # -- Ajuste 6: ICMS/Difal/FCP sempre positivos -------------------------
        icms  = abs(float(r.get('VALOR_ICMS',0)  or 0))
        difal = abs(float(r.get('VALOR_DIFAL',0) or 0))
        fcp   = abs(float(r.get('VALOR_FCP',0)   or 0))

        # -- Dados direto do arquivo_base (não depende mais do Cadastro de Clientes) --
        cnpj_cpf  = str(r.get('CPF_CNPJ_DESTINATARIO','')).strip()
        digitos   = ''.join(filter(str.isdigit, cnpj_cpf))
        cliente   = str(r.get('CLIENTE_DESTINATARIO','')).strip().upper()
        municipio = str(r.get('CIDADE_DESTINO','')).strip().upper()
        uf_dest   = str(r.get('UF_DESTINO','')).strip().upper()
        regiao    = get_regiao(uf_dest)

        # Tipo de Cliente:
        # - CPF (até 11 dígitos) -> sempre PESSOA FÍSICA
        # - CNPJ (14 dígitos) -> busca no cadastro; se não achar -> PESSOA JURÍDICA
        if len(digitos) <= 11:
            tipo_cli = 'PESSOA FÍSICA'
        else:
            tipo_cli = tipo_cliente_lookup(cnpj_cpf, cadastro_lookup) or 'PESSOA JURÍDICA'

        # -- Ajuste: NF como inteiro --------------------------------------------
        try:
            nf_val = int(float(r.get('NRO_NFE', 0)))
        except (ValueError, TypeError):
            nf_val = r.get('NRO_NFE', '')

        # Mapeamento EXATO seguindo a ordem das 44 colunas do Quadro de Envios
        # A=1 ... AQ=43 ... AR=44
        linha = {
            'Data Coleta'                : data_coleta,   # A  (1)  - só date
            'Ano'                        : data.year if pd.notna(data) else '',  # B (2)
            'Mês'                        : data.month if pd.notna(data) else '', # C (3)
            'CNPJ / CPF'                 : cnpj_cpf,      # D  (4)
            'Cliente'                    : cliente,        # E  (5)  - direto do Sankhya
            'País'                       : 'BRASIL',      # F  (6)
            'Municipio'                  : municipio,      # G  (7)  - direto do Sankhya
            'UF'                         : uf_dest,        # H  (8)  - direto do Sankhya
            'Região'                     : regiao,         # I  (9)  - calculado pela UF
            'Cep'                        : None,           # J  (10) - sempre vazio
            'Tipo de Cliente'            : tipo_cli,       # K  (11) - CPF=Física / Cadastro
            'No Pedido'                  : '',             # L  (12)
            'Cons Analise'               : 'Não',         # M  (13) - operador altera p/ Sim
            'Empresa'                    : empresa,        # N  (14)
            'NF'                         : nf_val,        # O  (15) - inteiro
            'NATUREZA DA OPERAÇÃO'       : natureza_raw,  # P  (16)
            'Cfop'                       : '',             # Q  (17)
            'Vr NF'                      : vr_nf,         # R  (18) - sempre positivo
            'Frete'                      : None,           # S  (19) - operador preenche
            'Peso (kg)'                  : abs(float(r.get('PESO',0) or 0)),     # T  (20)
            'Base Cálculo ICM'           : abs(float(r.get('VALOR_BASE_CALCULO',0) or 0)), # U (21)
            'ICMS'                       : icms,           # V  (22)
            'Difal'                     : difal,          # W  (23)
            'FCP'                        : fcp,            # X  (24)
            'Total Impostos'             : icms+difal+fcp, # Y  (25)
            'Volumes'                    : r.get('QTD_VOLUME', 0), # Z  (26)
            'Custo Embalagem'            : None,           # AA (27) - vazio
            'Transportadora'             : transp,         # AB (28)
            'Modalidade'                 : '',             # AC (29)
            'Prev. Envio'                : '',             # AD (30)
            'Prev. Entrega Cliente'      : '',             # AE (31)
            'Prev. Entrega Transporadora': '',             # AF (32)
            'DU'                         : '',             # AG (33)
            'Entrega Concluída?'         : 'Não',         # AH (34)
            'Data Efetiva da Entrega'    : '',             # AI (35)
            'Data Conf. Entrega'         : '',             # AJ (36)
            'Entrega No Prazo'           : '',             # AK (37) - vazio
            'Entrega em Atraso'          : '',             # AL (38) - vazio
            'Total Entregas'             : '',             # AM (39) - vazio
            'Origem da Venda'            : '',             # AN (40)
            'Cód Rastreio'               : '',             # AO (41)
            'Obs.'                       : '',             # AP (42)
            'Ano Ref'                    : ano_ref,        # AQ (43) - coluna correta!
            'Origem_Escola'              : '',             # AR (44)
        }
        # -- Replicar fórmulas ajustando número da linha ---------------------
        for col_idx, formula_base in formulas_ref.items():
            nova_formula = _re.sub(
                r'([A-Z]+)([0-9]+)',
                lambda m, r=prox_linha, ref=ultima_linha_ref: (
                    m.group(1) + str(int(m.group(2)) + (r - ref))
                ),
                formula_base
            )
            ws.cell(row=prox_linha, column=col_idx, value=nova_formula)

        # -- Escrever campos manuais pelo nome exato do cabeçalho --------------
        for campo, val in linha.items():
            if campo in CAMPOS_FORMULA:
                continue  # tem fórmula - não sobrescrever

            # Busca com e sem espaços (ex: 'Difal' e 'Difal' são o mesmo campo)
            col_idx = col_map.get(campo) or col_map.get(str(campo).strip())
            if col_idx is None:
                log(f"Campo '{campo}' não encontrado no cabeçalho", 'aviso')
                continue

            cell = ws.cell(row=prox_linha, column=col_idx, value=val)

            # NF: forçar inteiro numérico
            if campo == 'NF' and val not in ('', None):
                try:
                    cell.value         = int(float(val))
                    cell.data_type     = 'n'
                    cell.number_format = '0'
                except (ValueError, TypeError):
                    pass

            # Data Coleta: formato só data sem hora
            if campo == 'Data Coleta' and val is not None:
                cell.number_format = 'DD/MM/YYYY'

        # -- Transportadora vazia -> negrito ------------------------------------
        if not transp:
            col_idx_tr = col_map.get('Transportadora')
            if col_idx_tr:
                ws.cell(row=prox_linha, column=col_idx_tr).font = Font(bold=True)

        # -- Frete e CEP -> fundo cinza (operador deve preencher) --------------
        from openpyxl.styles import PatternFill
        cinza = PatternFill(fill_type='solid', fgColor='D9D9D9')
        for campo_cinza in ['Frete', 'Cep']:
            col_cinza = col_map.get(campo_cinza)
            if col_cinza:
                ws.cell(row=prox_linha, column=col_cinza).fill = cinza

        prox_linha += 1

    wb.save(QUADRO_ENVIOS)
    log(f"{len(df_novas)} NFs adicionadas ao Quadro de Envios!", 'ok')
    return True

# -- 3. PROCESSAMENTO DOS DADOS ------------------------------------------------
def norm(s):
    s = unicodedata.normalize('NFD', str(s).strip().upper())
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn')

def norm_nat(n):
    if not n or str(n).strip() in ('', 'None', 'nan'):
        return 'SEM NATUREZA'
    n = str(n).strip().upper()

    # VENDA (entram no Faturamento Bruto)
    if n in ('NOTA DE VENDA DE PRODUTO PRODUZIDO',
             'VENDA ENTREGA FUTURA - SIMPLES FAT'):
        return 'VENDA'

    # DEVOLUÇÃO que abate Faturamento Líquido
    if n in ('DEVOLUÇÃO DE VENDA - NF PRÓPRIA',
             'DEVOLUCAO DE VENDA - NF PROPRIA',
             'DEVOLUÇÃO DE VENDA - NF TERCEIROS',
             'DEVOLUCAO DE VENDA - NF TERCEIROS',
             'ESTORNO VDA - ENTREGA FUTURA SIMPLES FAT'):
        return 'DEVOLUÇÃO'

    # BONIFICAÇÃO (não abate)
    if n in ('REMESSA DE BONIFICAÇÃO - SAIDA',
             'REMESSA DE BONIFICACAO - SAIDA',
             'DEVOLUÇÃO DE REMESSA DE BONIFICAÇÃO',
             'DEVOLUCAO DE REMESSA DE BONIFICACAO',
             'DEVOLUÇÃO DE REMESSA - ENTREGA FUTURA',
             'DEVOLUCAO DE REMESSA - ENTREGA FUTURA'):
        return 'BONIFICAÇÃO'
    if 'BONIFIC' in n: return 'BONIFICAÇÃO'

    # SIMPLES REMESSA
    if n in ('SIMPLES REMESSA (OUTRAS)',):
        return 'SIMPLES REMESSA'

    # IGNORAR - não entram nos cálculos
    if n in ('AJUSTE DE ESTOQUE - SAIDA/DESCARTE',
             'AJUSTE DE ESTOQUE - SAÍDA DE ESTOQUE COM TERCEIROS',
             'AJUSTE DE ESTOQUE - SAIDA DE ESTOQUE COM TERCEIROS',
             'RETORNO SIMPLES REMESSA (OUTRAS) PROPRIO',
             'RETORNO SIMPLES REMESSA (OUTRAS) PRÓPRIO',
             'REMESSA VENDA ENTREGA FUTURA',
             'REMESSA EM CONSIGNAÇÃO - P/ VENDA',
             'REMESSA EM CONSIGNACAO - P/ VENDA',
             'VENDA CONSIGNADA',
             'TRANSFERENCIA ENTRE FILIAIS - SAIDA',
             'TRANSFERÊNCIA ENTRE FILIAIS - SAÍDA',
             'NOTA DE VENDA DE SERVIÇOS COALA',
             'NOTA DE VENDA DE SERVICOS COALA'):
        return 'IGNORAR'

    # Fallbacks genéricos
    if 'DEVOLU' in n and 'REMESSA' in n: return 'BONIFICAÇÃO'
    if 'DEVOLU' in n: return 'DEVOLUÇÃO'
    if n.startswith('VENDA') or 'NOTA DE VENDA' in n: return 'VENDA'
    if any(x in n for x in ('REMESSA','OUTRAS SAIDA','OUTRAS SAÍDA','USO FORA','ENVIO')):
        return 'SIMPLES REMESSA'
    if 'CANCELAD' in n or 'INUTILIZ' in n: return 'NF CANCELADA'
    return 'OUTROS'

def processar_dados():
    print("\n[3/3] PROCESSANDO DADOS -> ATUALIZANDO DASHBOARD")
    try:
        import pandas as pd
    except ImportError:
        log("pandas não instalado.", 'erro')
        return None

    # -- Carregar correlação e cadastro ----------------------------------------
    cor_map         = carregar_correlacao()
    cadastro_lookup = carregar_cadastro()

    log(f"Lendo: {QUADRO_ENVIOS.name}")
    df = pd.read_excel(QUADRO_ENVIOS, sheet_name='Lançamentos')

    # -- Filtro Cons Analise: APENAS "Sim" -------------------------------------
    # Ignora: "Não", "D", NaN, vazio - qualquer coisa que não seja "Sim"
    df = df[df['Cons Analise'] == 'Sim'].copy()
    log(f"Linhas válidas (Cons Analise=Sim): {len(df)}")

    # -- Normalizar campos -----------------------------------------------------
    # Normalizar colunas com espaços no nome (ex: 'Difal' -> 'Difal')
    df.columns = [c.strip() for c in df.columns]

    df['Empresa'] = df['Empresa'].fillna('').str.strip().str.upper().replace(
        {'VERBUM EDUCACAO S.A FILIAL': 'VERBUM', 'SERENA': 'IAB', 'Iab': 'IAB', 'iab': 'IAB', 'Verbum': 'VERBUM'})

    df['Transp_Norm'] = df['Transportadora'].apply(lambda n: map_transp(n, cor_map))
    df['Nat_Norm']    = df['NATUREZA DA OPERAÇÃO'].apply(norm_nat)
    df['Regiao']      = df['Região'].fillna('').str.strip().str.upper()
    df['Região']      = df['Regiao']  # alias
    df['Municipio']   = df['Municipio'].fillna('').str.strip().str.upper()
    df['UF']          = df['UF'].fillna('').str.strip().str.upper()

    # -- Normalizar nomes de Escolas e Municípios via tabela_correspondencia ----
    import re as _re_cnpj
    def _nc(c):
        return _re_cnpj.sub(r'[.\-/\s,]', '', str(c)).strip().lstrip('0') if c else ''

    _nome_map_global = {}
    try:
        _tab_path2 = PASTA_RAIZ / 'tabela_correspondencia.xlsx'
        if _tab_path2.exists():
            for _sht in ['Escolas', 'Municípios']:
                _df_tab = pd.read_excel(_tab_path2, sheet_name=_sht, header=2, engine='openpyxl')
                _df_tab.columns = [str(c).strip() for c in _df_tab.columns]
                _col_cnpj2 = next((c for c in _df_tab.columns if 'CNPJ' in c.upper()), None)
                _col_pad2  = next((c for c in _df_tab.columns if 'PADRONIZADO' in c.upper()), None)
                if _col_cnpj2 and _col_pad2:
                    for _, _r in _df_tab.iterrows():
                        _cnpj2 = _nc(str(_r[_col_cnpj2]))
                        _pad2  = str(_r[_col_pad2]).strip()
                        if _cnpj2 and _pad2 and _pad2.lower() not in ('nan','none',''):
                            _nome_map_global[_cnpj2] = _pad2
            log(f'Tabela correspondência carregada: {len(_nome_map_global)} nomes', 'ok')
        else:
            log('tabela_correspondencia.xlsx não encontrada na pasta raiz', 'aviso')
    except Exception as _e2:
        log(f'Erro ao carregar tabela_correspondencia: {_e2}', 'aviso')

    if _nome_map_global:
        df['Cliente'] = df.apply(
            lambda r: _nome_map_global.get(_nc(str(r['CNPJ / CPF'])), r['Cliente']),
            axis=1
        )

    # Detectar nome da coluna de data de coleta
    _col_data = next((c for c in df.columns if 'coleta' in c.lower()), 'Data Coleta')
    if _col_data != 'Data Coleta': df = df.rename(columns={_col_data: 'Data Coleta'})
    df['Data Coleta']             = pd.to_datetime(df['Data Coleta'], errors='coerce')
    df['Prev. Entrega Cliente']   = pd.to_datetime(df['Prev. Entrega Cliente'], errors='coerce')
    df['Data Efetiva da Entrega'] = pd.to_datetime(df['Data Efetiva da Entrega'], errors='coerce')
    df['Ano Ref']                 = pd.to_numeric(df['Ano Ref'], errors='coerce').fillna(0).astype(int)
    df['CNPJ / CPF']              = df['CNPJ / CPF'].astype(str).str.strip()
    df['NF']                      = df['NF'].astype(str).str.strip()

    df = df[df['Data Coleta'].dt.year >= 2000].copy()
    df['Data_Str'] = df['Data Coleta'].dt.strftime('%Y-%m-%d')
    df['AnoMes']   = df['Data Coleta'].dt.to_period('M').astype(str)

    for col in ['Vr NF','Frete','ICMS','Difal','FCP','Total Impostos','Peso (kg)','Custo Embalagem']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    df['Custo_Logistico'] = df['Frete'] + df['Custo Embalagem']

    # Enriquecer Tipo_Cliente via cadastro
    if cadastro_lookup:
        mask_vazio = df['Tipo de Cliente'].isna() | (df['Tipo de Cliente'].astype(str).str.strip() == '')
        if mask_vazio.any():
            df.loc[mask_vazio, 'Tipo de Cliente'] = df.loc[mask_vazio, 'CNPJ / CPF'].apply(
                lambda c: tipo_cliente_lookup(c, cadastro_lookup))
            log(f"Tipo_Cliente enriquecido: {mask_vazio.sum()} registros", 'ok')

    # -- Deduplicação de NFs ---------------------------------------------------
    df_dedup = df.drop_duplicates(subset=['NF','CNPJ / CPF'], keep='last').copy()

    # -- REGRA DATAS: efetiva <= coleta → usar Prev. Entrega Cliente -----------
    mask_dt_inv = (
        df_dedup['Data Efetiva da Entrega'].notna() &
        df_dedup['Data Coleta'].notna() &
        (df_dedup['Data Efetiva da Entrega'].dt.date <= df_dedup['Data Coleta'].dt.date)
    )
    if mask_dt_inv.any():
        df_dedup.loc[mask_dt_inv, 'Data Efetiva da Entrega'] = df_dedup.loc[mask_dt_inv, 'Prev. Entrega Cliente']
        log(f"Datas corrigidas (efetiva<=coleta): {mask_dt_inv.sum()}", 'ok')

    # -- CÁLCULO DE ENTREGAS - REGRA CORRETA -----------------------------------
    # 1 entrega física = mesmo Data Coleta + CNPJ/CPF + Cliente + Transportadora
    # Independente do número de NFs (NFs complementares não contam separado)
    # Atraso: Data Efetiva da Entrega > Prev. Entrega Cliente (por data, sem hora)
    # Data Efetiva: menor data do grupo (corrige erro de operador)
    # Entrega não concluída (sem Data Efetiva): NÃO contabilizada -
    #   será recalculada na próxima execução quando o operador preencher

    # Chave: mesmo cliente + mesma cidade + mesma transp normalizada = 1 entrega
    # Cidades diferentes com mesma transp = entregas distintas
    KEY_ENTREGA = ['Data_Str', 'CNPJ / CPF', 'Cliente', 'Transp_Norm', 'Municipio',
                   'Ano Ref', 'UF', 'Regiao', 'Empresa']

    # Separar concluídas (com data efetiva) das pendentes (sem data efetiva)
    df_concluidas = df_dedup[df_dedup['Data Efetiva da Entrega'].notna()].copy()
    df_pendentes  = df_dedup[df_dedup['Data Efetiva da Entrega'].isna()].copy()
    log(f"df_dedup: {len(df_dedup)} | concluídas: {len(df_concluidas)} | pendentes: {len(df_pendentes)}", 'ok')

    entregas = []
    sem_data_ef = len(df_pendentes)  # pendentes contadas diretamente
    for keys_val, g in df_concluidas.groupby(KEY_ENTREGA, dropna=False):
        datas_ef = g['Data Efetiva da Entrega'].dropna()

        if len(datas_ef) == 0:
            sem_data_ef += 1
            continue

        data_ef = datas_ef.min()  # menor data se houver divergência (erro operador)
        prevs   = g['Prev. Entrega Cliente'].dropna()
        prev_dt = prevs.min() if len(prevs) > 0 else None

        if prev_dt is not None:
            em_atraso = 1 if data_ef.date() > prev_dt.date() else 0
            no_prazo  = 1 - em_atraso
        else:
            em_atraso = 0
            no_prazo  = 0

        data_str, cnpj, cliente, transp, mun, ano, uf, reg, emp = keys_val
        try:
            ano_int = int(ano)
        except (ValueError, TypeError):
            continue  # ignorar linhas com Ano Ref inválido
        entregas.append({
            'Data_Str':  data_str,
            'Ano Ref':   ano_int,
            'Empresa':   str(emp),
            'Transp_Norm': transp,
            'Municipio': mun,
            'UF':        uf,
            'Regiao':    reg,
            'CNPJ / CPF': cnpj,
            'entregas':  1,
            'no_prazo':  no_prazo,
            'em_atraso': em_atraso,
            'Vr NF':     round(float(g['Vr NF'].sum()), 2),
            'Frete':     round(float(g['Frete'].sum()), 2),
            'Peso (kg)': round(float(g['Peso (kg)'].sum()), 3),
            'Volumes':   int(g['Volumes'].sum()),
            'AnoMes':    g['AnoMes'].iloc[0],
        })

    df_ent = pd.DataFrame(entregas) if entregas else pd.DataFrame()
    if len(df_ent) > 0 and 'sem_prev' not in df_ent.columns:
        df_ent['sem_prev'] = 0
    # df_ent_ns: apenas entregas COM previsão preenchida (para cálculo correto de NS/SLA)
    df_ent_ns = df_ent[df_ent['sem_prev'] == 0].copy() if len(df_ent) > 0 else pd.DataFrame()
    log(f"Entregas consolidadas: {len(df_ent)} | com previsão (NS): {len(df_ent_ns)} | Não concluídas: {sem_data_ef}", 'ok')

    # -- TRANSP_SLA e TRANSP_FULL para a aba Transportadoras ------------------
    transp_sla  = []
    transp_full = []

    if len(df_ent_ns) > 0:
        df_ent_2526 = df_ent_ns[df_ent_ns['Ano Ref'].isin([2025, 2026])].copy()

        for (ano, transp), g in df_ent_2526.groupby(['Ano Ref','Transp_Norm']):
            tot   = int(g['entregas'].sum())
            prazo = int(g['no_prazo'].sum())
            atr   = int(g['em_atraso'].sum())
            transp_sla.append({
                'ano': int(ano), 'transp': transp,
                'entregas': tot, 'no_prazo': prazo, 'em_atraso': atr,
                'sla':    round(prazo / max(tot, 1) * 100, 1),
                'frete':  round(float(g['Frete'].sum()), 2),
                'valor':  round(float(g['Vr NF'].sum()), 2),
                'peso':   round(float(g['Peso (kg)'].sum()), 3),
                'volumes':int(g['Volumes'].sum()),
            })

        for (ano, transp, mun, uf, reg), g in df_ent_2526.groupby(
                ['Ano Ref','Transp_Norm','Municipio','UF','Regiao']):
            transp_full.append({
                'ano': int(ano), 'transp': transp,
                'municipio': mun, 'uf': uf, 'regiao': reg,
                'valor':     round(float(g['Vr NF'].sum()), 2),
                'frete':     round(float(g['Frete'].sum()), 2),
                'peso':      round(float(g['Peso (kg)'].sum()), 3),
                'volumes':   int(g['Volumes'].sum()),
                'entregas':  int(g['entregas'].sum()),
                'no_prazo':  int(g['no_prazo'].sum()),
                'em_atraso': int(g['em_atraso'].sum()),
            })

        log(f"TRANSP_SLA: {len(transp_sla)} | TRANSP_FULL: {len(transp_full)}", 'ok')

    # -- RENT_FULL - rentabilidade por cliente (2025 em diante) ---------------
    df_rent_src = df_dedup[df_dedup['Ano Ref'] >= 2025].copy()
    rv = df_rent_src[df_rent_src['Nat_Norm']=='VENDA'].groupby(
        ['Ano Ref','Empresa','CNPJ / CPF','Cliente','Municipio','UF','Regiao','Tipo de Cliente']
    ).agg(envios=('NF','count'), vr_vendas=('Vr NF','sum'),
          icms=('ICMS','sum'), difal=('Difal','sum'), fcp=('FCP','sum'),
          total_imp=('Total Impostos','sum'), frete=('Frete','sum'),
          embalagem=('Custo Embalagem','sum'), custo_log=('Custo_Logistico','sum')).reset_index()
    rd = df_rent_src[df_rent_src['Nat_Norm']=='DEVOLUÇÃO'].groupby(
        ['Ano Ref','Empresa','CNPJ / CPF']).agg(vr_devolucoes=('Vr NF','sum')).reset_index()
    rent_full_df = rv.merge(rd, on=['Ano Ref','Empresa','CNPJ / CPF'], how='left')
    rent_full_df['vr_devolucoes'] = rent_full_df['vr_devolucoes'].fillna(0)
    rent_full_df['fat_liquido']   = rent_full_df['vr_vendas'] - rent_full_df['vr_devolucoes']
    rent_full_df['custo_total']   = rent_full_df['total_imp'] + rent_full_df['custo_log']
    rent_full_df['pct_custo_fat'] = (rent_full_df['custo_total'] /
                                     rent_full_df['vr_vendas'].replace(0,1)*100).round(2)
    rent_full_df['vr_nf']         = rent_full_df['vr_vendas']
    rent_full_df = rent_full_df.sort_values(['Empresa','vr_vendas'], ascending=[True,False]).round(2)
    rent_full_df = rent_full_df.rename(columns={
        'Ano Ref':'ano','CNPJ / CPF':'CNPJ','Tipo de Cliente':'Tipo_Cliente'
    })
    rent_full_list = rent_full_df.to_dict('records')
    log(f"RENT_FULL: {len(rent_full_list)} clientes", 'ok')

    # -- Função de KPIs por grupo (usa df_dedup para financeiro, df_ent para SLA) --
    def calc_kpis_grupo(d_fin, d_ent_grupo):
        tot_ent = int(d_ent_grupo['entregas'].sum()) if len(d_ent_grupo) > 0 else 0
        prazo   = int(d_ent_grupo['no_prazo'].sum())  if len(d_ent_grupo) > 0 else 0
        atraso  = int(d_ent_grupo['em_atraso'].sum()) if len(d_ent_grupo) > 0 else 0
        vr_dev  = round(float(d_fin[d_fin['Nat_Norm']=='DEVOLUÇÃO']['Vr NF'].sum()), 2)
        vr_vend = round(float(d_fin[d_fin['Nat_Norm']=='VENDA']['Vr NF'].sum()), 2)
        return {
            'total_nfs'      : int(len(d_fin)),
            'total_entregas' : tot_ent,
            'total_frete'    : round(float(d_fin['Frete'].sum()), 2),
            'total_embalagem': round(float(d_fin['Custo Embalagem'].sum()), 2),
            'total_custo_log': round(float(d_fin['Custo_Logistico'].sum()), 2),
            'total_vr_nf'    : round(float(d_fin['Vr NF'].sum()), 2),
            'total_icms'     : round(float(d_fin['ICMS'].sum()), 2),
            'total_difal'    : round(float(d_fin['Difal'].sum()), 2),
            'total_fcp'      : round(float(d_fin['FCP'].sum()), 2),
            'total_impostos' : round(float(d_fin['Total Impostos'].sum()), 2),
            'vr_devolucoes'  : vr_dev,
            'vr_vendas'      : vr_vend,
            'fat_liquido'    : round(vr_vend - vr_dev, 2),
            'clientes_unicos': int(d_fin['CNPJ / CPF'].nunique()),
            'entregues'      : tot_ent,
            'no_prazo'       : prazo,
            'em_atraso'      : atraso,
            'pct_prazo'      : round(prazo / max(tot_ent, 1) * 100, 1),
            'pct_custo_fat'  : round(float(d_fin['Custo_Logistico'].sum()) /
                                     max(float(d_fin['Vr NF'].sum()), 1) * 100, 2),
        }

    def calc_ano(d_fin, d_ent_grupo):
        kpis = calc_kpis_grupo(d_fin, d_ent_grupo)

        # Por mês
        pm_ent = d_ent_grupo.groupby('AnoMes').agg(entregas=('entregas','sum')).reset_index() \
                 if len(d_ent_grupo) > 0 else pd.DataFrame(columns=['AnoMes','entregas'])
        pm_fin = d_fin.groupby('AnoMes').agg(
            frete=('Frete','sum'), vr_nf=('Vr NF','sum'),
            icms=('ICMS','sum'), difal=('Difal','sum'), fcp=('FCP','sum'),
            custo_log=('Custo_Logistico','sum')
        ).reset_index()
        por_mes = pm_ent.merge(pm_fin, on='AnoMes', how='outer').sort_values('AnoMes').round(2)

        # Por transportadora (financeiro)
        pt_fin = d_fin.groupby('Transp_Norm').agg(
            frete=('Frete','sum'), peso=('Peso (kg)','sum')).reset_index()
        pt_ent = d_ent_grupo.groupby('Transp_Norm').agg(
            entregas=('entregas','sum')).reset_index() \
            if len(d_ent_grupo) > 0 else pd.DataFrame(columns=['Transp_Norm','entregas'])
        por_transp = pt_ent.merge(pt_fin, on='Transp_Norm', how='outer').fillna(0)
        por_transp['frete_medio'] = (por_transp['frete'] /
                                      por_transp['entregas'].replace(0,1)).round(2)
        por_transp = por_transp.sort_values('entregas', ascending=False).head(10).round(2)

        # SLA por transportadora
        if len(d_ent_grupo) > 0:
            tp_sla = d_ent_grupo.groupby('Transp_Norm').agg(
                entregas=('entregas','sum'),
                no_prazo=('no_prazo','sum'),
                em_atraso=('em_atraso','sum'),
            ).reset_index()
            tp_fin = d_fin.groupby('Transp_Norm').agg(frete=('Frete','sum')).reset_index()
            transp_prazo = tp_sla.merge(tp_fin, on='Transp_Norm', how='left')
            transp_prazo['pct_prazo'] = (
                transp_prazo['no_prazo'] /
                transp_prazo['entregas'].replace(0,1) * 100).round(1)
            transp_prazo = transp_prazo.sort_values('entregas', ascending=False).head(10).round(2)
        else:
            transp_prazo = pd.DataFrame()

        # Por natureza
        por_nat = d_fin.groupby('Nat_Norm').agg(
            entregas=('NF','count'), vr_nf=('Vr NF','sum'), frete=('Frete','sum')
        ).reset_index().sort_values('entregas', ascending=False).round(2)

        # Por região
        por_regiao = d_ent_grupo.groupby('Regiao').agg(
            entregas=('entregas','sum')).reset_index() \
            if len(d_ent_grupo) > 0 else pd.DataFrame(columns=['Regiao','entregas'])
        pr_fin = d_fin.groupby('Regiao').agg(
            frete=('Frete','sum'), custo_log=('Custo_Logistico','sum')).reset_index()
        por_regiao = por_regiao.merge(pr_fin, on='Regiao', how='outer').sort_values(
            'entregas', ascending=False).round(2)

        return {
            'kpis'        : kpis,
            'por_mes'     : por_mes.to_dict('records'),
            'por_transp'  : por_transp.to_dict('records'),
            'por_nat'     : por_nat.to_dict('records'),
            'por_regiao'  : por_regiao.to_dict('records'),
            'transp_prazo': transp_prazo.to_dict('records') if len(transp_prazo) > 0 else [],
        }

    # -- Coords IBGE -----------------------------------------------------------
    ibge_coords = {}
    UF_MAP = {
        11:'RO',12:'AC',13:'AM',14:'RR',15:'PA',16:'AP',17:'TO',
        21:'MA',22:'PI',23:'CE',24:'RN',25:'PB',26:'PE',27:'AL',28:'SE',29:'BA',
        31:'MG',32:'ES',33:'RJ',35:'SP',41:'PR',42:'SC',43:'RS',
        50:'MS',51:'MT',52:'GO',53:'DF'
    }
    if MUNICIPIOS_JSON.exists():
        with open(MUNICIPIOS_JSON, encoding='utf-8-sig') as f:
            muns = json.load(f)
        for m in muns:
            uf_code = int(str(m['codigo_ibge'])[:2])
            uf = UF_MAP.get(uf_code, '')
            key = norm(m['nome']) + '|' + uf
            ibge_coords[key] = {
                'cod_ibge': str(m['codigo_ibge']),
                'lat': m['latitude'], 'lon': m['longitude']
            }
        log(f"Coords IBGE: {len(ibge_coords)} municípios")

    # -- Loop por ano ----------------------------------------------------------
    anos_ref      = sorted(df_dedup['Ano Ref'].unique().tolist())
    dados_por_ano = {}
    municipios_total = []

    for ano in anos_ref:
        if ano == 0: continue
        d_fin      = df_dedup[df_dedup['Ano Ref'] == ano].copy()
        d_ent_ano  = df_ent[df_ent['Ano Ref'] == ano].copy() if len(df_ent) > 0 else pd.DataFrame()

        dados_total = calc_ano(d_fin, d_ent_ano)

        # Por empresa
        dados_emp    = {}
        sla_emp      = []
        por_emp_list = []

        for emp in ['IAB', 'VERBUM']:
            de      = d_fin[d_fin['Empresa'] == emp]
            de_ent  = d_ent_ano[d_ent_ano['Empresa'] == emp] if len(d_ent_ano) > 0 else pd.DataFrame()
            if len(de) == 0: continue
            dados_emp[emp] = calc_ano(de, de_ent)
            np_ = int(de_ent['no_prazo'].sum())  if len(de_ent) > 0 else 0
            at_ = int(de_ent['em_atraso'].sum()) if len(de_ent) > 0 else 0
            ent = int(de_ent['entregas'].sum())  if len(de_ent) > 0 else 0
            sla_emp.append({'Empresa': emp, 'entregues': ent, 'no_prazo': np_,
                            'em_atraso': at_,
                            'pct_prazo': round(np_ / max(ent, 1) * 100, 1)})

        for emp in d_fin['Empresa'].unique():
            de = d_fin[d_fin['Empresa'] == emp]
            por_emp_list.append({
                'Empresa'      : emp,
                'entregas'     : int(len(de)),
                'frete'        : round(float(de['Frete'].sum()), 2),
                'vr_nf'        : round(float(de['Vr NF'].sum()), 2),
                'icms'         : round(float(de['ICMS'].sum()), 2),
                'difal'        : round(float(de['Difal'].sum()), 2),
                'fcp'          : round(float(de['FCP'].sum()), 2),
                'embalagem'    : round(float(de['Custo Embalagem'].sum()), 2),
                'custo_log'    : round(float(de['Custo_Logistico'].sum()), 2),
                'vr_vendas'    : round(float(de[de['Nat_Norm']=='VENDA']['Vr NF'].sum()), 2),
                'vr_devolucoes': round(float(de[de['Nat_Norm']=='DEVOLUÇÃO']['Vr NF'].sum()), 2),
                'fat_liquido'  : round(
                    float(de[de['Nat_Norm']=='VENDA']['Vr NF'].sum()) -
                    float(de[de['Nat_Norm']=='DEVOLUÇÃO']['Vr NF'].sum()), 2),
            })

        mapa_uf = d_fin.groupby(['UF','Empresa']).agg(
            clientes=('CNPJ / CPF','nunique'), entregas=('NF','count'),
            vr_nf=('Vr NF','sum'), frete=('Frete','sum')
        ).reset_index().round(2)

        # Rentabilidade
        rv   = d_fin[d_fin['Nat_Norm']=='VENDA'].groupby(
            ['Empresa','CNPJ / CPF','Cliente','Municipio','UF','Regiao','Tipo de Cliente']
        ).agg(envios=('NF','count'), vr_vendas=('Vr NF','sum'),
              icms=('ICMS','sum'), difal=('Difal','sum'), fcp=('FCP','sum'),
              total_imp=('Total Impostos','sum'), frete=('Frete','sum'),
              embalagem=('Custo Embalagem','sum'), custo_log=('Custo_Logistico','sum')).reset_index()
        rd   = d_fin[d_fin['Nat_Norm']=='DEVOLUÇÃO'].groupby(
            ['Empresa','CNPJ / CPF']).agg(vr_devolucoes=('Vr NF','sum')).reset_index()
        rent = rv.merge(rd, on=['Empresa','CNPJ / CPF'], how='left')
        rent['vr_devolucoes'] = rent['vr_devolucoes'].fillna(0)
        rent['fat_liquido']   = rent['vr_vendas'] - rent['vr_devolucoes']
        rent['custo_total']   = rent['total_imp'] + rent['custo_log']
        rent['pct_custo_fat'] = (rent['custo_total'] / rent['vr_vendas'].replace(0,1)*100).round(2)
        rent = rent.sort_values(['Empresa','vr_vendas'], ascending=[True,False]).round(2)

        # Municípios com coords
        grp_mun    = d_fin.groupby(['Municipio','UF','Empresa']).agg(
            entregas=('NF','count'), clientes=('CNPJ / CPF','nunique'), vr_nf=('Vr NF','sum')
        ).reset_index()
        municipios = []
        for _, row in grp_mun.iterrows():
            key  = norm(row['Municipio']) + '|' + row['UF']
            info = ibge_coords.get(key)
            if info:
                m = {
                    'cod_ibge' : info['cod_ibge'],
                    'municipio': row['Municipio'],
                    'uf'       : row['UF'],
                    'empresa'  : row['Empresa'],
                    'ano_ref'  : ano,
                    'entregas' : int(row['entregas']),
                    'clientes' : int(row['clientes']),
                    'vr_nf'    : round(float(row['vr_nf']), 2),
                    'lat'      : info['lat'],
                    'lon'      : info['lon'],
                }
                municipios.append(m)
                municipios_total.append(m)

        dados_por_ano[str(ano)] = {
            **dados_total,
            'por_emp'      : por_emp_list,
            'sla_emp'      : sla_emp,
            'mapa_uf'      : mapa_uf.to_dict('records'),
            'rentabilidade': rent.to_dict('records'),
            'dados_emp'    : dados_emp,
            'municipios'   : municipios,
        }
        log(f"Ano {ano}: {dados_total['kpis']['total_entregas']} entregas | "
            f"SLA {dados_total['kpis']['pct_prazo']}% | "
            f"Fat. líquido R${dados_total['kpis']['fat_liquido']:,.0f}", 'ok')

    # Carrega geo anterior
    geo_ant = None
    if JSON_DASHBOARD.exists():
        try:
            with open(JSON_DASHBOARD, encoding='utf-8') as f:
                ant = json.load(f)
            geo_ant = ant.get('geo_municipios')
        except Exception:
            pass

    # -- NS por empresa x ano (para atualizar H2.ns_iab e H2.VERBUM.ns) ---------
    ns_iab_por_ano    = {}
    ns_verbum_por_ano = {}
    if len(df_ent_ns) > 0:
        for (ano, emp), g in df_ent_ns.groupby(['Ano Ref', 'Empresa']):
            prazo = int(g['no_prazo'].sum())
            total = int(g['entregas'].sum())
            ns    = round(prazo / max(total, 1) * 100, 1)
            if str(emp).upper() == 'IAB':
                ns_iab_por_ano[int(ano)] = ns
            elif str(emp).upper() == 'VERBUM':
                ns_verbum_por_ano[int(ano)] = ns

    # ── Calcular campos por empresa para DRE completo ─────────────────────────
    fat_venda_por_ano = {}; devol_por_ano = {}
    frete_tot = {}; emb_tot = {}; icms_tot = {}; difal_tot = {}; fcp_tot = {}
    iab_por_ano = {}; verbum_por_ano = {}

    for ano in anos_ref:
        # Total por ano — usa df_dedup (dedupado por NF+CNPJ, Cons=Sim, 2025/2026)
        g_all = df_dedup[df_dedup['Ano Ref']==ano] if df_dedup is not None else None
        if g_all is not None:
            fat_venda_por_ano[ano] = round(float(g_all[g_all['Nat_Norm']=='VENDA']['Vr NF'].sum()), 2)
            devol_por_ano[ano]     = round(float(g_all[g_all['Nat_Norm']=='DEVOLUÇÃO']['Vr NF'].sum()), 2)
            frete_tot[ano]  = round(float(g_all['Frete'].sum()), 2)
            emb_tot[ano]    = round(float(g_all['Custo Embalagem'].sum()) if 'Custo Embalagem' in g_all else 0, 2)
            icms_tot[ano]   = round(float(g_all['ICMS'].sum()), 2)
            difal_tot[ano]  = round(float(g_all['Difal'].sum()), 2)
            fcp_tot[ano]    = round(float(g_all['FCP'].sum()), 2)
            # IAB
            g_iab = g_all[g_all['Empresa']=='IAB']
            iab_por_ano[ano] = {
                'fat':   round(float(g_iab[g_iab['Nat_Norm']=='VENDA']['Vr NF'].sum()), 2),
                'frete': round(float(g_iab['Frete'].sum()), 2),
                'devol': round(float(g_iab[g_iab['Nat_Norm']=='DEVOLUÇÃO']['Vr NF'].sum()), 2),
                'icms':  round(float(g_iab['ICMS'].sum()), 2),
                'difal': round(float(g_iab['Difal'].sum()), 2),
                'fcp':   round(float(g_iab['FCP'].sum()), 2),
                'emb':   round(float(g_iab['Custo Embalagem'].sum()) if 'Custo Embalagem' in g_iab else 0, 2),
            }
            # VERBUM
            g_vbm = g_all[g_all['Empresa']=='VERBUM']
            g_ent_vbm = df_ent_ns[(df_ent_ns['Ano Ref']==ano) & (df_ent_ns['Empresa']=='VERBUM')] if len(df_ent_ns) > 0 else None
            ent_vbm = int(g_ent_vbm['entregas'].sum()) if g_ent_vbm is not None and len(g_ent_vbm) > 0 else 0
            praz_vbm= int(g_ent_vbm['no_prazo'].sum())  if g_ent_vbm is not None and len(g_ent_vbm) > 0 else 0
            ns_vbm  = round(praz_vbm/max(ent_vbm,1)*100,1) if ent_vbm > 0 else None
            verbum_por_ano[ano] = {
                'fat':   round(float(g_vbm[g_vbm['Nat_Norm']=='VENDA']['Vr NF'].sum()), 2),
                'frete': round(float(g_vbm['Frete'].sum()), 2),
                'devol': round(float(g_vbm[g_vbm['Nat_Norm']=='DEVOLUÇÃO']['Vr NF'].sum()), 2),
                'icms':  round(float(g_vbm['ICMS'].sum()), 2),
                'difal': round(float(g_vbm['Difal'].sum()), 2),
                'fcp':   round(float(g_vbm['FCP'].sum()), 2),
                'emb':   round(float(g_vbm['Custo Embalagem'].sum()) if 'Custo Embalagem' in g_vbm else 0, 2),
                'ent':   ent_vbm,
                'ns':    ns_vbm,
            }

    # ── Aba Acompanhamento: Pendentes e Atrasadas ─────────────────────────────
    from datetime import date as _dt_date, timedelta as _timedelta
    import re as _norm_re

    def _norm_cnpj_local(c):
        return _norm_re.sub(r'[.\-/\s,]', '', str(c)).strip().lstrip('0') if c else ''

    # Carregar mapa de nomes padronizados
    _nome_map = {}
    try:
        _tab_path = PASTA_RAIZ / 'tabela_correspondencia.xlsx'
        if _tab_path.exists():
            import openpyxl as _opx
            _wb = _opx.load_workbook(_tab_path, read_only=True, data_only=True)
            for _sht in ['Escolas', 'Municípios']:
                if _sht in _wb.sheetnames:
                    _ws = _wb[_sht]
                    _hdrs = [str(c.value).strip() if c.value else '' for c in next(_ws.iter_rows(min_row=3, max_row=3))]
                    _col_pad = next((i for i,h in enumerate(_hdrs) if 'PADRONIZADO' in h.upper()), None)
                    _col_cnpj= next((i for i,h in enumerate(_hdrs) if 'CNPJ' in h.upper()), 0)
                    if _col_pad is not None:
                        for _row in _ws.iter_rows(min_row=4, values_only=True):
                            _cnpj = _norm_cnpj_local(str(_row[_col_cnpj]) if _row[_col_cnpj] else '')
                            _pad  = str(_row[_col_pad]).strip() if _row[_col_pad] else ''
                            if _pad and _pad.lower() not in ('none','nan',''): _nome_map[_cnpj] = _pad
    except Exception as _e:
        log(f'tabela_correspondencia não carregada: {_e}', 'aviso')

    def _nome_pad(cnpj, nome):
        return _nome_map.get(_norm_cnpj_local(str(cnpj)), str(nome))

    def _fmt_d(d):
        try: return pd.Timestamp(d).strftime('%d/%m/%Y') if pd.notna(d) else ''
        except: return ''
    def _fmt_nf(v):
        try: return str(int(float(str(v).split('.')[0])))
        except: return str(v)
    def _clean(v):
        s = str(v) if v is not None else ''; return '' if s.lower() in ('nan','none','') else s

    # df para aba acompanhamento — todos anos, Cons=Sim
    _df_ac = df_dedup.copy()
    _df_ac['Prev. Entrega Cliente']   = pd.to_datetime(_df_ac['Prev. Entrega Cliente'],   errors='coerce')
    _df_ac['Data Efetiva da Entrega'] = pd.to_datetime(_df_ac['Data Efetiva da Entrega'], errors='coerce')
    _df_ac['Data Coleta']             = pd.to_datetime(_df_ac['Data Coleta'],             errors='coerce')
    _df_ac['Cliente_Pad'] = _df_ac.apply(lambda r: _nome_pad(r['CNPJ / CPF'], r['Cliente']), axis=1)

    _hoje = _dt_date.today()
    _KEY_AC = ['Data Coleta','CNPJ / CPF','Transp_Norm','Municipio','UF','Ano Ref']

    # Quadro 1: Pendentes (Cons=Sim, Não concluída, sem data efetiva, vence em <=2 dias)
    _df_pend = _df_ac[
        (_df_ac['Entrega Concluída?'].astype(str).str.strip() == 'Não') &
        (_df_ac['Data Efetiva da Entrega'].isna()) &
        (_df_ac['Prev. Entrega Cliente'].notna()) &
        (_df_ac['Prev. Entrega Cliente'].dt.date <= _hoje + _timedelta(days=2))
    ].copy().sort_values('Prev. Entrega Cliente')

    pend_list = []
    for _kv, _g in _df_pend.groupby(_KEY_AC, dropna=False):
        _prev = _g['Prev. Entrega Cliente'].dropna().min()
        _dias = (_prev.date() - _hoje).days if pd.notna(_prev) else None
        _nfs  = sorted(set(_fmt_nf(n) for n in _g['NF'] if _fmt_nf(n) not in ('nan','None','')))
        _obs  = [_clean(o) for o in _g['Obs.'] if _clean(o)]
        _dt_col, _cnpj, _tr, _mun, _uf, _ano = _kv
        pend_list.append({
            'ano': int(float(str(_ano))) if str(_ano) not in ('nan','') else 0,
            'nfs': ', '.join(_nfs) if _nfs else '—',
            'data_coleta': _fmt_d(_dt_col),
            'cliente': _clean(_g['Cliente_Pad'].iloc[0]),
            'municipio': _clean(_mun), 'uf': _clean(_uf), 'transp': _clean(_tr),
            'prev_entrega': _fmt_d(_prev), 'obs': ' | '.join(_obs),
            'dias_rest': _dias,
            'status': 'VENCE HOJE' if _dias==0 else ('ATRASADA' if _dias and _dias<0 else f'{_dias}d'),
        })

    # Quadro 2: Concluídas em atraso
    _df_atr = _df_ac[
        (_df_ac['Entrega Concluída?'].astype(str).str.strip() == 'Sim') &
        (_df_ac['Data Efetiva da Entrega'].notna()) &
        (_df_ac['Prev. Entrega Cliente'].notna()) &
        (_df_ac['Data Efetiva da Entrega'].dt.date > _df_ac['Prev. Entrega Cliente'].dt.date)
    ].copy().sort_values('Data Efetiva da Entrega', ascending=False)

    atr_list = []
    _KEY_ATR = ['Data Coleta','CNPJ / CPF','Transp_Norm','Municipio','UF','Ano Ref','Data Efetiva da Entrega']
    for _kv, _g in _df_atr.groupby(_KEY_AC, dropna=False):
        _prev = _g['Prev. Entrega Cliente'].dropna().min()
        _ef   = _g['Data Efetiva da Entrega'].dropna().min()
        _dias_atr = (_ef.date()-_prev.date()).days if pd.notna(_prev) and pd.notna(_ef) else 0
        _nfs  = sorted(set(_fmt_nf(n) for n in _g['NF'] if _fmt_nf(n) not in ('nan','None','')))
        _obs  = [_clean(o) for o in _g['Obs.'] if _clean(o)]
        _dt_col, _cnpj, _tr, _mun, _uf, _ano = _kv
        atr_list.append({
            'ano': int(float(str(_ano))) if str(_ano) not in ('nan','') else 0,
            'nfs': ', '.join(_nfs) if _nfs else '—',
            'data_coleta': _fmt_d(_dt_col),
            'cliente': _clean(_g['Cliente_Pad'].iloc[0]),
            'municipio': _clean(_mun), 'uf': _clean(_uf), 'transp': _clean(_tr),
            'prev_entrega': _fmt_d(_prev), 'data_efetiva': _fmt_d(_ef),
            'dias_atraso': _dias_atr, 'obs': ' | '.join(_obs),
        })

    log(f'Acompanhamento: {len(pend_list)} pendentes | {len(atr_list)} atrasadas', 'ok')

    resultado = {
        'anos'             : anos_ref,
        'dados'            : dados_por_ano,
        'heatmap'          : municipios_total,
        'municipios'       : municipios_total,
        'transp_sla'       : transp_sla,
        'transp_full'      : transp_full,
        'rent_full'        : rent_full_list,
        'ns_iab_por_ano'   : ns_iab_por_ano,
        'ns_verbum_por_ano': ns_verbum_por_ano,
        'fat_venda_por_ano': fat_venda_por_ano,
        'devol_por_ano'    : devol_por_ano,
        'frete_total_2025' : frete_tot.get(2025, 0),
        'frete_total_2026' : frete_tot.get(2026, 0),
        'emb_total_2025'   : emb_tot.get(2025, 0),
        'emb_total_2026'   : emb_tot.get(2026, 0),
        'icms_total_2025'  : icms_tot.get(2025, 0),
        'icms_total_2026'  : icms_tot.get(2026, 0),
        'difal_total_2025' : difal_tot.get(2025, 0),
        'difal_total_2026' : difal_tot.get(2026, 0),
        'fcp_total_2025'   : fcp_tot.get(2025, 0),
        'fcp_total_2026'   : fcp_tot.get(2026, 0),
        'iab_por_ano'      : iab_por_ano,
        'verbum_por_ano'   : verbum_por_ano,
        'pend_list'        : pend_list,
        'atr_list'         : atr_list,
    }
    if geo_ant:
        resultado['geo_municipios'] = geo_ant

    with open(JSON_DASHBOARD, 'w', encoding='utf-8') as f:
        json.dump(resultado, f, ensure_ascii=False, default=str)
    log(f"JSON salvo: {JSON_DASHBOARD.name}", 'ok')
    return resultado


# -- HISTÓRICO H2 --------------------------------------------------------------
def calcular_historico():
    """Calcula série histórica 2012-atual e retorna bloco JS H2."""
    print("\n[HIST] CALCULANDO SÉRIE HISTÓRICA")
    try:
        import pandas as pd
        from collections import defaultdict
    except ImportError:
        log("pandas não instalado", 'erro')
        return None

    cadastro_lookup = carregar_cadastro()
    cor_map         = carregar_correlacao()

    def calcular_ano_hist(g):
        """KPIs de um ano dos arquivos históricos."""
        g = g[g['CONS ANÁLISE'].astype(str).str.strip().str.upper() != 'D'].copy()
        for col in ['VR NF','CUSTO DO FRETE RATEADO']:
            g[col] = pd.to_numeric(g[col], errors='coerce').fillna(0)
        g['DATA DE COLETA']              = pd.to_datetime(g['DATA DE COLETA'], errors='coerce')
        g['DATA EFETIVA DA ENTREGA']     = pd.to_datetime(g['DATA EFETIVA DA ENTREGA'], errors='coerce')
        g['PREVISAO DE ENTREGA CLIENTE'] = pd.to_datetime(g['PREVISAO DE ENTREGA CLIENTE'], errors='coerce')

        fat   = round(float(g['VR NF'].sum()), 2)
        frete = round(float(g['CUSTO DO FRETE RATEADO'].sum()), 2)
        uf_vol   = {str(k):int(v) for k,v in g.groupby('UF').size().items()}
        tr_frete = {str(k):round(float(v),2) for k,v in
                    g.groupby('TRANSP')['CUSTO DO FRETE RATEADO'].sum().items()}

        col_tipo = next((c for c in g.columns
                         if 'TIPO' in str(c).upper() and 'CLIENTE' in str(c).upper()), None)
        if col_tipo:
            tipo_vol = {str(k):int(v) for k,v in g.groupby(col_tipo).size().items()}
        elif cadastro_lookup and 'CNPJ / CPF' in g.columns:
            g['_Tipo'] = g['CNPJ / CPF'].apply(
                lambda c: tipo_cliente_lookup(c, cadastro_lookup) or 'PESSOA FÍSICA')
            tipo_vol = {str(k):int(v) for k,v in g.groupby('_Tipo').size().items()}
        else:
            tipo_vol = {}

        # -- Consolidação de entregas (mesma regra do Quadro) ------------------
        cnpj_col = 'CNPJ / CPF' if 'CNPJ / CPF' in g.columns else \
                   next((c for c in g.columns if 'CNPJ' in str(c).upper()), None)
        client_col = 'CLIENTE' if 'CLIENTE' in g.columns else \
                     next((c for c in g.columns if 'CLIENT' in str(c).upper()), None)

        if cnpj_col and client_col:
            envios = g.groupby(
                ['DATA DE COLETA', cnpj_col, client_col, 'TRANSP'], dropna=False
            ).agg(
                data_ef=('DATA EFETIVA DA ENTREGA', lambda x: x.dropna().min()),
                prev_dt=('PREVISAO DE ENTREGA CLIENTE', lambda x: x.dropna().min()),
            ).reset_index()
            entregues = envios[envios['data_ef'].notna()]
        else:
            envios    = g.groupby(['DATA DE COLETA','TRANSP'], dropna=False).agg(
                data_ef=('DATA EFETIVA DA ENTREGA', lambda x: x.dropna().min()),
                prev_dt=('PREVISAO DE ENTREGA CLIENTE', lambda x: x.dropna().min()),
            ).reset_index()
            entregues = envios[envios['data_ef'].notna()]

        total  = len(entregues)
        com_pv = entregues[entregues['prev_dt'].notna()]
        prazo  = int((com_pv['data_ef'].dt.date <= com_pv['prev_dt'].dt.date).sum())
        atraso = int((com_pv['data_ef'].dt.date >  com_pv['prev_dt'].dt.date).sum())
        ns     = round(prazo/(prazo+atraso)*100,1) if (prazo+atraso)>0 else None

        tr_prazo  = {}
        tr_atraso = {}
        for tr, grp in entregues.groupby('TRANSP'):
            cp = grp[grp['prev_dt'].notna()]
            tr_prazo[str(tr)]  = int((cp['data_ef'].dt.date <= cp['prev_dt'].dt.date).sum())
            tr_atraso[str(tr)] = int((cp['data_ef'].dt.date >  cp['prev_dt'].dt.date).sum())

        return dict(fat=fat, frete=frete, entregas=total, prazo=prazo, atraso=atraso,
                    ns=ns, uf_vol=uf_vol, tr_frete=tr_frete,
                    tr_prazo=tr_prazo, tr_atraso=tr_atraso, tipo_vol=tipo_vol)

    def calcular_ano_quadro(df_dedup, df_ent, ano):
        """KPIs de um ano do Quadro de Envios - usa entregas consolidadas."""
        d_fin     = df_dedup[df_dedup['Ano Ref'] == ano]
        d_ent_ano = df_ent[df_ent['Ano Ref'] == ano] if len(df_ent) > 0 else pd.DataFrame()

        fat   = round(float(d_fin['Vr NF'].sum()), 2)
        frete = round(float(d_fin['Frete'].sum()), 2)
        ent   = int(d_ent_ano['entregas'].sum()) if len(d_ent_ano) > 0 else 0
        np_   = int(d_ent_ano['no_prazo'].sum())  if len(d_ent_ano) > 0 else 0
        at_   = int(d_ent_ano['em_atraso'].sum()) if len(d_ent_ano) > 0 else 0
        ns    = round(np_/max(np_+at_,1)*100,1)

        uf_vol    = {str(k):int(v) for k,v in d_fin.groupby('UF').size().items()}
        tr_frete  = {str(k):round(float(v),2)
                     for k,v in d_fin.groupby('Transp_Norm')['Frete'].sum().items()}
        tr_prazo  = {}
        tr_atraso = {}
        if len(d_ent_ano) > 0:
            for tr, g in d_ent_ano.groupby('Transp_Norm'):
                tr_prazo[str(tr)]  = int(g['no_prazo'].sum())
                tr_atraso[str(tr)] = int(g['em_atraso'].sum())
        tipo_vol  = {str(k):int(v) for k,v in
                     d_fin.groupby('Tipo de Cliente').size().items() if pd.notna(k)}

        return dict(fat=fat, frete=frete, entregas=ent, prazo=np_, atraso=at_,
                    ns=ns, uf_vol=uf_vol, tr_frete=tr_frete,
                    tr_prazo=tr_prazo, tr_atraso=tr_atraso, tipo_vol=tipo_vol)

    kpis = {}

    # Históricos
    for fpath, anos in [(HIST_2012,[2012,2013,2014,2015]),
                        (HIST_2016,[2016,2017,2018,2019]),
                        (HIST_2020,[2020,2021,2022,2023,2024])]:
        if fpath is None:
            log(f"Histórico não encontrado - pulando {anos}", 'aviso')
            continue
        log(f"Lendo {fpath.name}...")
        try:
            df = pd.read_excel(fpath, sheet_name='Lançamentos', usecols=range(39))
            df['ANO REF'] = pd.to_numeric(df['ANO REF'], errors='coerce')
            for ano in anos:
                g = df[df['ANO REF']==ano].copy()
                if len(g) == 0: continue
                kpis[ano] = calcular_ano_hist(g)
                log(f"  {ano}: {kpis[ano]['entregas']} entregas | NS {kpis[ano]['ns']}%", 'ok')
        except Exception as e:
            log(f"Erro em {fpath.name}: {e}", 'erro')

    # Quadro de Envios (2025+) - reprocessa com regra de consolidação
    if QUADRO_ENVIOS.exists():
        log(f"Lendo {QUADRO_ENVIOS.name}...")
        try:
            df_q = pd.read_excel(QUADRO_ENVIOS, sheet_name='Lançamentos')
            df_q = df_q[df_q['Cons Analise'] == 'Sim'].copy()

            for col in ['Vr NF','Frete']:
                df_q[col] = pd.to_numeric(df_q[col], errors='coerce').fillna(0)
            df_q['Data Coleta']             = pd.to_datetime(df_q['Data Coleta'], errors='coerce')
            df_q['Prev. Entrega Cliente']   = pd.to_datetime(df_q['Prev. Entrega Cliente'], errors='coerce')
            df_q['Data Efetiva da Entrega'] = pd.to_datetime(df_q['Data Efetiva da Entrega'], errors='coerce')
            df_q['Ano Ref']     = pd.to_numeric(df_q['Ano Ref'], errors='coerce').fillna(0).astype(int)
            df_q['CNPJ / CPF']  = df_q['CNPJ / CPF'].astype(str).str.strip()
            df_q['Transp_Norm'] = df_q['Transportadora'].apply(lambda n: map_transp(n, cor_map))
            df_q['Data_Str']    = df_q['Data Coleta'].dt.strftime('%Y-%m-%d')
            df_q['Tipo de Cliente'] = df_q['Tipo de Cliente'].fillna('')
            df_q['Custo Embalagem'] = pd.to_numeric(df_q['Custo Embalagem'], errors='coerce').fillna(0)
            df_q['Custo_Logistico'] = df_q['Frete'] + df_q['Custo Embalagem']
            df_q = df_q[df_q['Data Coleta'].dt.year >= 2000]
            df_dedup = df_q.drop_duplicates(subset=['NF','CNPJ / CPF'], keep='last').copy()

            # Consolidar entregas com a regra correta
            KEY = ['Data_Str','CNPJ / CPF','Cliente','Transp_Norm','Ano Ref',
                   'Municipio','UF','Região','Empresa']
            ent_list = []
            for keys_val, g in df_dedup.groupby(KEY, dropna=False):
                datas_ef = g['Data Efetiva da Entrega'].dropna()
                if len(datas_ef) == 0: continue
                data_ef = datas_ef.min()
                prevs   = g['Prev. Entrega Cliente'].dropna()
                prev_dt = prevs.min() if len(prevs) > 0 else None
                if prev_dt is not None:
                    em_atraso = 1 if data_ef.date() > prev_dt.date() else 0
                    no_prazo  = 1 - em_atraso
                else:
                    em_atraso = 0; no_prazo = 0
                ent_list.append({
                    'Ano Ref':     keys_val[4],
                    'Transp_Norm': keys_val[3],
                    'Empresa':     keys_val[8],
                    'no_prazo':    no_prazo,
                    'em_atraso':   em_atraso,
                    'entregas':    1,
                })
            df_ent = pd.DataFrame(ent_list) if ent_list else pd.DataFrame()

            for ano in sorted(df_dedup['Ano Ref'].unique()):
                if ano == 0: continue
                kpis[ano] = calcular_ano_quadro(df_dedup, df_ent, ano)
                log(f"  {ano}: {kpis[ano]['entregas']} entregas | NS {kpis[ano]['ns']}%", 'ok')
        except Exception as e:
            log(f"Erro ao ler Quadro de Envios: {e}", 'erro')

    if not kpis:
        log("Nenhum dado histórico encontrado.", 'aviso')
        return None

    from collections import defaultdict
    import json as _json

    anos     = sorted(kpis.keys())
    fat_js   = [round(float(kpis[a]['fat']),2)   for a in anos]
    frete_js = [round(float(kpis[a]['frete']),2) for a in anos]
    ent_js   = [int(kpis[a]['entregas'])          for a in anos]
    ns_js    = [float(kpis[a]['ns']) if kpis[a]['ns'] else None for a in anos]

    uf_total = defaultdict(int)
    for d in kpis.values():
        for uf, v in d['uf_vol'].items():
            if isinstance(uf, str) and len(uf) == 2:
                uf_total[uf] += int(v)
    uf15 = sorted(uf_total.items(), key=lambda x:-x[1])[:15]
    uf_k = [u for u,v in uf15]
    uf_v = [int(v) for u,v in uf15]

    tr_p = defaultdict(int); tr_a = defaultdict(int); tr_ft = defaultdict(float)
    for d in kpis.values():
        for tr,v in d['tr_prazo'].items():
            if isinstance(tr,str) and tr.strip(): tr_p[tr] += int(v)
        for tr,v in d['tr_atraso'].items():
            if isinstance(tr,str) and tr.strip(): tr_a[tr] += int(v)
        for tr,v in d['tr_frete'].items():
            if isinstance(tr,str) and tr.strip(): tr_ft[tr] += float(v)

    tr_envios = {tr: tr_p[tr]+tr_a[tr]
                 for tr in set(list(tr_p)+list(tr_a)) if tr.strip()}
    top10     = sorted(tr_envios.items(), key=lambda x:-x[1])[:10]
    tr_objs   = []
    for tr, e in top10:
        p=tr_p[tr]; a=tr_a[tr]; ft=round(float(tr_ft[tr]),2)
        fm=round(ft/e,2) if e>0 else 0
        sla=round(p/(p+a)*100,1) if (p+a)>0 else 0
        tr_objs.append(f"    {{n:'{tr}',e:{e},ft:{ft},fm:{fm},p:{p},a:{a},s:{sla}}}")

    tot_ent  = sum(int(kpis[a]['entregas']) for a in anos)
    tot_fat  = round(sum(float(kpis[a]['fat'])   for a in anos), 2)
    tot_fret = round(sum(float(kpis[a]['frete']) for a in anos), 2)
    all_p    = sum(int(kpis[a]['prazo'])  for a in anos)
    all_a    = sum(int(kpis[a]['atraso']) for a in anos)
    ns_med   = round(all_p/(all_p+all_a)*100,1) if (all_p+all_a)>0 else 0

    inicio_ano = min(anos); fim_ano = max(anos)

    JS = f"""// ===== HISTÓRICO {inicio_ano}-{fim_ano} (ANO REF) =====
var H2 = {{
  labels: {_json.dumps([str(a) for a in anos])},
  fat:    {_json.dumps(fat_js)},
  frete:  {_json.dumps(frete_js)},
  ent:    {_json.dumps(ent_js)},
  ns:     {_json.dumps(ns_js)},
  ufK:    {_json.dumps(uf_k)},
  ufV:    {_json.dumps(uf_v)},
  transp: [
{chr(10).join(tr_objs)}
  ],
  totEnt:   {tot_ent},
  totFat:   {tot_fat},
  totFrete: {tot_fret},
  nsMed:    {ns_med}
}};"""

    log(f"Histórico: {len(anos)} anos ({inicio_ano}-{fim_ano}) | {tot_ent:,} entregas", 'ok')
    return JS


# -- ATUALIZAR HTML ------------------------------------------------------------
def atualizar_html(raw):
    """Atualiza TODOS os dados do dashboard HTML a partir dos dados processados."""
    import re as _re
    from collections import defaultdict as _dd
    from datetime import date as _date, timedelta as _td

    log(f"Atualizando {HTML_DASHBOARD.name}...", 'proc')
    with open(HTML_DASHBOARD, encoding='utf-8') as f:
        html = f.read()

    # ── 1. RAW ────────────────────────────────────────────────────────────────
    js_data = json.dumps(raw, ensure_ascii=False)
    html = _re.sub(r'const RAW=\{.*?\};', f'const RAW={js_data};', html, flags=_re.DOTALL)

    # ── 2. HMP_DADOS ──────────────────────────────────────────────────────────
    if 'var HMP_DADOS' in html and raw.get('municipios'):
        js_hmp = json.dumps(raw['municipios'], ensure_ascii=False)
        html = _re.sub(r'var HMP_DADOS = \[.*?\];', f'var HMP_DADOS = {js_hmp};',
                       html, flags=_re.DOTALL)

    # ── 3. TRANSP_SLA, TRANSP_FULL, RENT_FULL ────────────────────────────────
    for const, key in [('TRANSP_FULL','transp_full'),('TRANSP_SLA','transp_sla'),('RENT_FULL','rent_full')]:
        if not raw.get(key): continue
        marker = f'const {const} = '
        idx = html.find(marker)
        if idx >= 0:
            end = html.index(';\n', idx)
            html = html[:idx] + marker + json.dumps(raw[key], ensure_ascii=False) + ';\n' + html[end+2:]
            log(f'{const}: {len(raw[key])} registros', 'ok')

    # ── 4. H2 — atualizar todos os campos ────────────────────────────────────
    if 'var H2 = {' not in html:
        log('H2 não encontrado no HTML', 'erro')
        return

    # Calcular by_ano a partir do TRANSP_SLA
    by_ano = _dd(lambda: {'ent':0,'prazo':0,'atr':0})
    for r in raw.get('transp_sla',[]):
        by_ano[r['ano']]['ent']   += r['entregas']
        by_ano[r['ano']]['prazo'] += r['no_prazo']
        by_ano[r['ano']]['atr']   += r['em_atraso']

    idx_h2 = html.find('var H2 = {'); end_h2 = html.find('\n};', idx_h2)+3
    h2 = html[idx_h2:end_h2]

    # Labels
    lab_m = _re.search(r'labels:\s+\[([^\]]+)\]', h2)
    if not lab_m:
        log('labels não encontrado no H2', 'erro')
        return
    labels = [x.strip().strip('"') for x in lab_m.group(1).split(',')]
    ANOS = [int(l) for l in labels]

    # Função auxiliar para atualizar índices 2025/2026 num array
    def upd(arr, val_2025, val_2026):
        a = list(arr)
        if 2025 in ANOS: a[ANOS.index(2025)] = val_2025
        if 2026 in ANOS: a[ANOS.index(2026)] = val_2026
        return a

    # 4a. H2.ent e H2.ns
    ent_m = _re.search(r'ent:\s+\[([^\]]+)\]', h2)
    ns_m  = _re.search(r'ns:\s+\[([^\]]+)\]',  h2)
    if ent_m and ns_m:
        ent_vals = [int(float(x.strip())) for x in ent_m.group(1).split(',')]
        ns_vals  = [float(x.strip()) if x.strip() not in ('null','None') else None
                    for x in ns_m.group(1).split(',')]
        for ano_k, d in by_ano.items():
            if ano_k in ANOS:
                i = ANOS.index(ano_k)
                ent_vals[i] = d['ent']
                ns_vals[i]  = round(d['prazo']/max(d['ent'],1)*100, 1)
        h2 = _re.sub(r'ent:\s+\[[^\]]+\]', f'ent:    {json.dumps(ent_vals)}', h2)
        ns_js = '[' + ','.join(str(v) if v is not None else 'null' for v in ns_vals) + ']'
        h2 = _re.sub(r'ns:\s+\[[^\]]+\]',  f'ns:     {ns_js}', h2)

    # 4b. H2.ns_iab
    ns_iab_map    = raw.get('ns_iab_por_ano', {})
    ns_verbum_map = raw.get('ns_verbum_por_ano', {})
    if ns_iab_map and ns_m:
        ns_iab_arr = list(ns_vals)
        for i, ano in enumerate(ANOS):
            if ano in ns_iab_map: ns_iab_arr[i] = ns_iab_map[ano]
        ns_iab_js = '[' + ','.join(str(v) if v is not None else 'null' for v in ns_iab_arr) + ']' 
        h2 = _re.sub(r'ns_iab:\s*\[[^\]]+\]', f'ns_iab: {ns_iab_js}', h2)

    # 4c. porNat VENDA e DEVOLUÇÃO — atualizar 2025/2026 com valores reais
    for nat_key, raw_key in [('VENDA','fat_venda_por_ano'),('DEVOLU','devol_por_ano')]:
        nat_vals = raw.get(raw_key, {})
        if not nat_vals: continue
        pat = rf'("{nat_key}[^"]*":\s*\{{"todos":\s*\{{"fat":\[)([^\]]+)\]'
        m = _re.search(pat, h2)
        if m:
            arr = [float(x.strip()) for x in m.group(2).split(',')]
            for ano, val in nat_vals.items():
                if int(ano) in ANOS: arr[ANOS.index(int(ano))] = val
            h2 = h2[:m.start()] + m.group(1) + ','.join(str(round(v,2)) for v in arr) + ']' + h2[m.end():]

    # 4d. H2.frete, emb, icms, difal, fcp — atualizar 2025/2026
    for campo, raw_key_25, raw_key_26 in [
        ('frete', raw.get('frete_total_2025',0), raw.get('frete_total_2026',0)),
        ('emb',   raw.get('emb_total_2025',0),   raw.get('emb_total_2026',0)),
        ('icms',  raw.get('icms_total_2025',0),  raw.get('icms_total_2026',0)),
        ('difal', raw.get('difal_total_2025',0), raw.get('difal_total_2026',0)),
        ('fcp',   raw.get('fcp_total_2025',0),   raw.get('fcp_total_2026',0)),
    ]:
        pat = rf'{campo}:\s+\[([^\]]+)\]'
        m = _re.search(pat, h2)
        if m:
            arr = [float(x.strip()) for x in m.group(1).split(',')]
            arr = upd(arr, raw_key_25, raw_key_26)
            h2 = h2[:m.start()] + f'{campo}:   {json.dumps([round(v,2) for v in arr])}' + h2[m.end():]

    # 4e. H2.VERBUM completo
    vbm = raw.get('verbum_por_ano', {})
    if vbm:
        idx_v = h2.find('VERBUM: {')
        if idx_v >= 0:
            depth=0; pos=idx_v+8
            while pos < len(h2):
                if h2[pos]=='{': depth+=1
                elif h2[pos]=='}':
                    depth-=1
                    if depth==0: break
                pos+=1
            verbum_old = h2[idx_v:pos+1]

            def build_arr(campo):
                arr = [0.0]*len(ANOS)
                for ano, d in vbm.items():
                    if int(ano) in ANOS: arr[ANOS.index(int(ano))] = d.get(campo, 0)
                return arr

            vbm_fat   = build_arr('fat')
            vbm_frete = build_arr('frete')
            vbm_devol = build_arr('devol')
            vbm_emb   = build_arr('emb')
            vbm_icms  = build_arr('icms')
            vbm_difal = build_arr('difal')
            vbm_fcp   = build_arr('fcp')
            vbm_ent   = build_arr('ent')
            vbm_ns    = [vbm.get(a, vbm.get(str(a), {})).get('ns', None) for a in ANOS]
            # fallback: build_arr usa int(ano), vbm pode ter chaves int ou str
            if all(v is None for v in vbm_ns):
                vbm_ns = [next((d.get('ns') for k,d in vbm.items() if int(k)==a), None) for a in ANOS]
            vbm_ns_js = '['+','.join(str(v) if v is not None else 'null' for v in vbm_ns)+']' 

            verbum_new = (f'VERBUM: {{fat:{json.dumps([round(v,2) for v in vbm_fat])},'
                          f'frete:{json.dumps([round(v,2) for v in vbm_frete])},'
                          f'devol:{json.dumps([round(v,2) for v in vbm_devol])},'
                          f'emb:{json.dumps([round(v,2) for v in vbm_emb])},'
                          f'icms:{json.dumps([round(v,2) for v in vbm_icms])},'
                          f'difal:{json.dumps([round(v,2) for v in vbm_difal])},'
                          f'fcp:{json.dumps([round(v,2) for v in vbm_fcp])},'
                          f'ent:{json.dumps([int(v) for v in vbm_ent])},'
                          f'ns:{vbm_ns_js}}}'  )
            h2 = h2[:idx_v] + verbum_new + h2[pos+1:]
            log('H2.VERBUM completo atualizado', 'ok')

    # 4f. porTipo2 IAB fat e frete — atualizar 2025/2026
    iab_por_ano = raw.get('iab_por_ano', {})
    if iab_por_ano:
        iab_fat_m = _re.search(r'("IAB":\{"todos":\{"fat":\[)([^\]]+)\]', h2)
        if iab_fat_m:
            arr = [float(x.strip()) for x in iab_fat_m.group(2).split(',')]
            for ano, d in iab_por_ano.items():
                if int(ano) in ANOS: arr[ANOS.index(int(ano))] = d.get('fat', 0)
            h2 = h2[:iab_fat_m.start()] + iab_fat_m.group(1) + ','.join(str(round(v,2)) for v in arr) + ']' + h2[iab_fat_m.end():]

        iab_frt_m = _re.search(r'("IAB":\{"todos":\{"fat":\[[^\]]+\],"frete":\[)([^\]]+)\]', h2)
        if iab_frt_m:
            arr = [float(x.strip()) for x in iab_frt_m.group(2).split(',')]
            for ano, d in iab_por_ano.items():
                if int(ano) in ANOS: arr[ANOS.index(int(ano))] = d.get('frete', 0)
            h2 = h2[:iab_frt_m.start()] + iab_frt_m.group(1) + ','.join(str(round(v,2)) for v in arr) + ']' + h2[iab_frt_m.end():]

    # 4g. porTipo2 todos ent — sincronizar 2025/2026
    for ano_k, ent_val in [(2025, by_ano[2025]['ent']),(2026, by_ano[2026]['ent'])]:
        if ano_k not in ANOS: continue
        idx_l = ANOS.index(ano_k)
        def fix_ent(m, il=idx_l, ev=ent_val):
            vals = m.group(2).split(',')
            if il < len(vals): vals[il] = str(ev)
            return m.group(1) + ','.join(vals) + ']' 
        h2 = _re.sub(r'("ent":\[)([^\]]+)\]', fix_ent, h2)

    html = html[:idx_h2] + h2 + html[end_h2:]
    log('H2 sincronizado', 'ok')

    # ── 5. Aba Acompanhamento: PEND_DATA e ATR_DATA ───────────────────────────
    pend_list = raw.get('pend_list', [])
    atr_list  = raw.get('atr_list', [])

    while '<!-- Dados das Entregas Pendentes -->' in html:
        idx_s = html.find('<!-- Dados das Entregas Pendentes -->')
        idx_e = html.find('</script>', idx_s) + len('</script>')
        html  = html[:idx_s] + html[idx_e:]

    pos_func   = html.find('function filtrarPendencias')
    pos_script = html.rfind('<script>', 0, pos_func)
    novo = (f'<!-- Dados das Entregas Pendentes -->\n<script>\n'
            f'var PEND_DATA = {json.dumps(pend_list, ensure_ascii=False)};\n'
            f'var ATR_DATA  = {json.dumps(atr_list,  ensure_ascii=False)};\n'
            f'</script>\n')
    html = html[:pos_script] + novo + html[pos_script:]
    log(f'Acompanhamento: {len(pend_list)} pendentes | {len(atr_list)} atrasadas', 'ok')

    # ── 6. Sincronizar selects de ANO com anos reais dos dados ───────────────
    import re as _re2
    anos_disp   = sorted(set(r['ano'] for r in raw.get('transp_sla',[])))
    anos_labels = [str(a) for a in range(2012, max(anos_disp, default=2026)+1)]
    anos_rec    = [str(a) for a in anos_disp]  # só anos com dados reais

    def _mk_opts(anos, selected=None, com_todos=False, label_todos='Todos'):
        o = f'<option value="todos">{{label_todos}}</option>' if com_todos else ''
        for a in anos:
            sel = ' selected' if str(a)==str(selected) else ''
            o += f'<option value="{{a}}"{{sel}}>{{a}}</option>'
        return o

    def _fix_sel(h, sel_id, opts_html):
        pat = rf'(id="{{sel_id}}"[^>]*>)(.*?)(</select>)'
        m = _re2.search(pat, h, _re2.DOTALL)
        if m: h = h[:m.start()] + m.group(1) + opts_html + m.group(3) + h[m.end():]
        return h

    # compAnoInicio / compAnoFim — todos os anos históricos
    html = _fix_sel(html, 'compAnoInicio', _mk_opts(anos_labels, anos_labels[0]))
    html = _fix_sel(html, 'compAnoFim',    _mk_opts(anos_labels, anos_labels[-1]))
    # anoSel (Visão Executiva) — todos + anos recentes
    html = _fix_sel(html, 'anoSel',    _mk_opts(anos_rec, None, com_todos=True))
    # pendAnoSel — anos recentes + Todos, default último ano
    html = _fix_sel(html, 'pendAnoSel', _mk_opts(anos_rec, anos_rec[-1] if anos_rec else None, com_todos=True))
    # rentAnoSel — anos recentes + Todos
    html = _fix_sel(html, 'rentAnoSel', _mk_opts(anos_rec, None, com_todos=True))
    # hmpAnoSel — anos recentes + Todos
    html = _fix_sel(html, 'hmpAnoSel',  _mk_opts(anos_rec, None, com_todos=True))
    # transpAnoSel — anos recentes + Todos (2025+2026)
    label_tr = ' + '.join(anos_rec) if anos_rec else 'Todos'
    # Nota: transpAnoSel é um <select> na aba Transportadoras
    html = _fix_sel(html, 'transpAnoSel', _mk_opts(anos_rec, anos_rec[-1] if anos_rec else None,
                                                    com_todos=True, label_todos=label_tr))
    # Garantir que o transpAnoSel fechou corretamente (sem `>` faltando)
    html = html.replace('</select\n    </div>\n    <div style="overflow-x:auto;max-height:420px',
                        '</select>\n    </div>\n    <div style="overflow-x:auto;max-height:420px', 1)
    # anoSelOp (Operações) — anos recentes + Todos
    html = _fix_sel(html, 'anoSelOp', _mk_opts(anos_rec, None, com_todos=True))
    log(f'Selects de ano atualizados: {anos_rec}', 'ok')

    # ── 7. Gravar ─────────────────────────────────────────────────────────────
    with open(HTML_DASHBOARD, 'w', encoding='utf-8') as f:
        f.write(html)
    log(f'Dashboard salvo ({len(html)//1024} KB)', 'ok')

    # ── 8. Atualizar dashboard_neutro com os mesmos dados ────────────────────
    if HTML_NEUTRO.exists():
        with open(HTML_NEUTRO, encoding='utf-8') as f:
            html_neutro = f.read()

        # Aplicar os mesmos blocos de dados no neutro
        import re as _rn
        html_neutro = _rn.sub(r'const RAW=\{.*?\};', f'const RAW={json.dumps(raw, ensure_ascii=False)};', html_neutro, flags=_rn.DOTALL)

        for const, key in [('TRANSP_FULL','transp_full'),('TRANSP_SLA','transp_sla'),('RENT_FULL','rent_full')]:
            if not raw.get(key): continue
            marker = f'const {const} = '
            idx = html_neutro.find(marker)
            if idx >= 0:
                end = html_neutro.index(';\n', idx)
                html_neutro = html_neutro[:idx] + marker + json.dumps(raw[key], ensure_ascii=False) + ';\n' + html_neutro[end+2:]

        # Copiar bloco H2 já atualizado do dashboard_logistica
        idx_h2_orig = html.find('var H2 = {')
        end_h2_orig = html.find('\n};', idx_h2_orig) + 3
        h2_atualizado = html[idx_h2_orig:end_h2_orig]

        idx_h2_n = html_neutro.find('var H2 = {')
        end_h2_n = html_neutro.find('\n};', idx_h2_n) + 3
        if idx_h2_n >= 0:
            html_neutro = html_neutro[:idx_h2_n] + h2_atualizado + html_neutro[end_h2_n:]

        # Copiar PEND_DATA e ATR_DATA
        pend_list = raw.get('pend_list', [])
        atr_list  = raw.get('atr_list', [])
        while '<!-- Dados das Entregas Pendentes -->' in html_neutro:
            idx_s = html_neutro.find('<!-- Dados das Entregas Pendentes -->')
            idx_e = html_neutro.find('</script>', idx_s) + len('</script>')
            html_neutro = html_neutro[:idx_s] + html_neutro[idx_e:]
        pos_func   = html_neutro.find('function filtrarPendencias')
        pos_script = html_neutro.rfind('<script>', 0, pos_func)
        novo = (f'<!-- Dados das Entregas Pendentes -->\n<script>\n'
                f'var PEND_DATA = {json.dumps(pend_list, ensure_ascii=False)};\n'
                f'var ATR_DATA  = {json.dumps(atr_list,  ensure_ascii=False)};\n'
                f'</script>\n')
        html_neutro = html_neutro[:pos_script] + novo + html_neutro[pos_script:]

        with open(HTML_NEUTRO, 'w', encoding='utf-8') as f:
            f.write(html_neutro)
        log(f'Dashboard neutro salvo ({len(html_neutro)//1024} KB)', 'ok')
    else:
        log(f'dashboard_neutro.html não encontrado - pulando', 'aviso')


# ── COMERCIAL — Notion ────────────────────────────────────────────────────────
NOTION_TOKEN  = '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'
NOTION_DB_ID  = '27051690-649a-804c-84c5-d13455af4136'
JSON_COMERCIAL = PASTA_GITHUB / 'dados_comercial.json'

# Status que devem ser CONSIDERADOS no dashboard
STATUS_CONSIDERAR = {'05 - contrato enviado', '08 - material enviado'}
# Mapa de estados para regiões
REGIOES = {
    'AC':'Norte','AM':'Norte','AP':'Norte','PA':'Norte','RO':'Norte','RR':'Norte','TO':'Norte',
    'AL':'Nordeste','BA':'Nordeste','CE':'Nordeste','MA':'Nordeste','PB':'Nordeste',
    'PE':'Nordeste','PI':'Nordeste','RN':'Nordeste','SE':'Nordeste',
    'DF':'Centro-Oeste','GO':'Centro-Oeste','MS':'Centro-Oeste','MT':'Centro-Oeste',
    'ES':'Sudeste','MG':'Sudeste','RJ':'Sudeste','SP':'Sudeste',
    'PR':'Sul','RS':'Sul','SC':'Sul',
}

def _notion_get_prop(props, key):
    p = props.get(key)
    if not p: return ''
    t = p.get('type','')
    if t == 'title':       return ''.join(x.get('plain_text','') for x in p.get('title',[]))
    if t == 'rich_text':   return ''.join(x.get('plain_text','') for x in p.get('rich_text',[]))
    if t == 'select':      return (p.get('select') or {}).get('name','')
    if t == 'status':      return (p.get('status') or {}).get('name','')
    if t == 'multi_select':return ', '.join(s.get('name','') for s in p.get('multi_select',[]))
    if t == 'number':      return p.get('number') or 0
    if t == 'formula':
        f = p.get('formula',{})
        return f.get('number') or f.get('string','')
    if t == 'rollup':
        r = p.get('rollup',{})
        return r.get('number') or 0
    return ''

def _notion_find_key(props, *hints):
    keys = list(props.keys())
    for h in hints:
        found = next((k for k in keys if h.lower() in k.lower()), None)
        if found: return found
    return ''

def atualizar_comercial():
    """Busca dados do Notion e injeta JSON estático nos dashboards."""
    import urllib.request, urllib.error

    log('Buscando dados do Comercial no Notion...', 'proc')

    pages, cursor = [], None
    while True:
        body = json.dumps({'page_size': 100, **({'start_cursor': cursor} if cursor else {})}).encode()
        req  = urllib.request.Request(
            f'https://api.notion.com/v1/databases/{NOTION_DB_ID}/query',
            data=body,
            headers={
                'Authorization': f'Bearer {NOTION_TOKEN}',
                'Notion-Version': '2022-06-28',
                'Content-Type': 'application/json',
            },
            method='POST'
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as r:
                data = json.loads(r.read())
        except urllib.error.HTTPError as e:
            log(f'Erro HTTP ao acessar Notion: {e.code} {e.reason}', 'erro')
            return
        except Exception as e:
            log(f'Erro ao acessar Notion: {e}', 'erro')
            return

        pages.extend(data.get('results', []))
        if data.get('has_more'):
            cursor = data.get('next_cursor')
        else:
            break

    log(f'{len(pages)} registros encontrados no Notion', 'ok')

    # Debug: mostrar o campo Status raw da primeira página
    if pages:
        pr0 = pages[0].get('properties', {})
        status_raw = pr0.get('Status', {})
        log(f'Status raw (primeira página): {status_raw}', 'info')

    # Mostrar quais status existem no Notion (para debug)
    status_encontrados = set()
    for page in pages:
        pr = page.get('properties', {})
        s = str(_notion_get_prop(pr, 'Status')).strip()
        if s: status_encontrados.add(s)
    log(f'Status encontrados: {sorted(status_encontrados)}', 'info')

    # Mapear páginas
    registros = []
    for page in pages:
        pr = page.get('properties', {})
        fk = lambda *h: _notion_find_key(pr, *h)

        valor_raw = _notion_get_prop(pr, 'Propostas 2025/26 (R$)')
        try:
            if isinstance(valor_raw, (int, float)):
                valor = float(valor_raw)
            else:
                v = str(valor_raw).replace('.','').replace(',','.').strip()
                valor = float(v) if v else 0
        except:
            valor = 0

        if valor <= 0:
            continue

        status = str(_notion_get_prop(pr, 'Status')).strip()
        if not any(s in status.lower() for s in STATUS_CONSIDERAR):
            continue

        estado = str(_notion_get_prop(pr, 'Estado')).strip().upper()
        registros.append({
            'vendedor':    str(_notion_get_prop(pr, 'Vendedor')).strip(),
            'perfil':      str(_notion_get_prop(pr, 'PERFIL')).strip(),
            'cliente':     str(_notion_get_prop(pr, 'Cliente')).strip(),
            'estado':      estado,
            'municipio':   str(_notion_get_prop(pr, 'Município')).strip(),
            'valor':       round(valor, 2),
            'status':      status,
            'temperatura': str(_notion_get_prop(pr, 'Temperatura')).strip(),
            'tipoVenda':   str(_notion_get_prop(pr, 'Tipo de Venda')).strip(),
            'regiao':      REGIOES.get(estado, 'Outros'),
        })

    log(f'{len(registros)} registros considerados (com valor + status válido)', 'ok')

    # Calcular KPIs e agrupamentos
    valor_total   = sum(r['valor'] for r in registros)
    por_status    = {}
    por_regiao    = {}
    for r in registros:
        s = r['status']
        por_status[s] = por_status.get(s, 0) + r['valor']
        rg = r['regiao']
        por_regiao[rg] = por_regiao.get(rg, 0) + r['valor']

    payload = {
        'atualizado_em': datetime.now().strftime('%d/%m/%Y %H:%M'),
        'total':         len(registros),
        'valor_total':   round(valor_total, 2),
        'por_status':    por_status,
        'por_regiao':    por_regiao,
        'registros':     registros,
    }

    # Salvar JSON local (para debug/backup)
    with open(JSON_COMERCIAL, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    log(f'dados_comercial.json salvo ({len(registros)} registros)', 'ok')

    # Injetar nos dashboards HTML
    json_str = f'const COMERCIAL_DATA = {json.dumps(payload, ensure_ascii=False)};'
    marker_start = '/* COMERCIAL_DATA_START */'
    marker_end   = '/* COMERCIAL_DATA_END */'
    bloco = f'{marker_start}\n{json_str}\n{marker_end}'

    for html_file in [HTML_DASHBOARD, HTML_NEUTRO]:
        if not html_file.exists():
            continue
        with open(html_file, encoding='utf-8') as f:
            html = f.read()

        if marker_start in html:
            # Substituir bloco existente
            idx_s = html.find(marker_start)
            idx_e = html.find(marker_end) + len(marker_end)
            html  = html[:idx_s] + bloco + html[idx_e:]
        else:
            # Inserir antes do </body>
            html = html.replace('</body>', f'<script>\n{bloco}\n</script>\n</body>', 1)

        with open(html_file, 'w', encoding='utf-8') as f:
            f.write(html)
        log(f'{html_file.name} atualizado com dados do Comercial', 'ok')

    log('Comercial atualizado com sucesso!', 'ok')


if __name__ == '__main__':
    inicio = datetime.now()

    # -- Log automático em arquivo (pasta logs/ junto ao script) ---------------
    try:
        import pathlib as _pl
        _log_dir = _pl.Path(__file__).parent / 'logs'
        _log_dir.mkdir(parents=True, exist_ok=True)
        _log_file = _log_dir / f"dashboard_{inicio.strftime('%Y%m%d')}.txt"
        _log_fh = open(_log_file, 'a', encoding='utf-8')
        _log_fh.write(f"\n[{inicio.strftime('%d/%m/%Y %H:%M:%S')}] Iniciando\n")
        class _Tee:
            def __init__(self, *s): self.s = s
            def write(self, d):
                for x in self.s:
                    try: x.write(d)
                    except: pass
            def flush(self):
                for x in self.s:
                    try: x.flush()
                    except: pass
        sys.stdout = _Tee(sys.stdout, _log_fh)
        sys.stderr = _Tee(sys.stderr, _log_fh)
    except Exception as _le:
        print(f"Aviso: log em arquivo não iniciado — {_le}")

    # -- Modo de execução ------------------------------------------------------
    so_quadro    = '--so-quadro'    in sys.argv
    so_dashboard = '--so-dashboard' in sys.argv
    modo_completo = not so_quadro and not so_dashboard

    if so_quadro:
        titulo = "ATUALIZA QUADRO DE ENVIOS - LOGÍSTICA SERENA"
    elif so_dashboard:
        titulo = "ATUALIZA DASHBOARD - LOGÍSTICA SERENA"
    else:
        titulo = "ATUALIZAÇÃO COMPLETA - LOGÍSTICA SERENA"

    print("=" * 55)
    print(f"  {titulo}")
    print(f"  {inicio.strftime('%d/%m/%Y %H:%M')}")
    print("=" * 55)

    if so_quadro:
        # Só atualiza o Quadro de Envios
        if not fazer_backup():
            sys.exit(1)
        integrar_sankhya()

    elif so_dashboard:
        # Só atualiza o Dashboard
        raw = processar_dados()
        if raw is None:
            sys.exit(1)
        atualizar_html(raw)
        atualizar_comercial()

    else:
        # Modo completo - faz tudo
        if not fazer_backup():
            sys.exit(1)
        integrar_sankhya()
        raw = processar_dados()
        if raw is None:
            sys.exit(1)
        atualizar_html(raw)
        atualizar_comercial()

    fim = datetime.now()
    print(f"\n{'=' * 55}")
    print(f"  [OK] CONCLUÍDO em {(fim-inicio).seconds}s")
    if so_dashboard or modo_completo:
        print(f"  Abra o dashboard_logistica.html no navegador.")
    print(f"{'=' * 55}")
